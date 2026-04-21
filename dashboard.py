import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from datetime import timedelta
import sys
import os
import time
import requests
import re
from io import BytesIO
from supabase import create_client
import socket
from openpyxl.utils import get_column_letter

# Add the current directory to path
sys.path.append(os.path.dirname(__file__))

# Import authentication functions
from auth import show_login_page, show_profile_page, show_admin_panel

# ---------------------------------------------------
# Supabase Configuration - Using Streamlit Secrets
# ---------------------------------------------------
@st.cache_resource
def init_supabase():
    """Initialize Supabase client"""
    try:
        supabase_url = st.secrets["SUPABASE_URL"]
        supabase_key = st.secrets["SUPABASE_KEY"]
        supabase = create_client(supabase_url, supabase_key)
        return supabase
    except Exception as e:
        st.error(f"Error connecting to Supabase: {e}")
        return None

def check_supabase_connection():
    """Check Supabase connection health"""
    try:
        if st.session_state.supabase_client:
            response = st.session_state.supabase_client.table("health_data").select("*").limit(1).execute()
            return True
    except Exception:
        return False
    return False

# Check authentication
if 'auth' not in st.session_state:
    st.session_state['auth'] = False

if not st.session_state['auth']:
    show_login_page()
    st.stop()

# ---------------------------------------------------
# Page Setup
# ---------------------------------------------------
st.set_page_config(
    page_title="Health Program Medicines Dashboard", 
    layout="wide", 
    initial_sidebar_state="expanded"
)

# ---------------------------------------------------
# CSS Styling with Mobile Responsiveness (ENHANCED)
# ---------------------------------------------------
st.markdown("""
<style>
    /* Main container */
    .stApp {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
    }

    /* Metric cards */
    div[data-testid="stMetric"] {
        background: white;
        border-radius: 15px;
        padding: 15px;
        transition: all 0.3s ease;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }

    div[data-testid="stMetric"]:hover {
        transform: translateY(-5px);
        box-shadow: 0 8px 15px rgba(0,0,0,0.2);
        background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
    }

    /* Animations */
    @keyframes shake {
        0%, 100% { transform: translateX(0); }
        10%, 30%, 50%, 70%, 90% { transform: translateX(-2px); }
        20%, 40%, 60%, 80% { transform: translateX(2px); }
    }

    @keyframes vertical-shake {
        0%, 100% { transform: translateY(0); }
        10%, 30%, 50%, 70%, 90% { transform: translateY(-2px); }
        20%, 40%, 60%, 80% { transform: translateY(2px); }
    }

    @keyframes gentle-shake {
        0%, 100% { transform: translateX(0); }
        25%, 75% { transform: translateX(-1px); }
        50% { transform: translateX(1px); }
    }

    .stButton button:hover {
        animation: shake 0.5s ease-in-out;
        transform: scale(1.02);
        transition: all 0.3s ease;
    }

    div[data-testid="stMetricValue"]:hover {
        animation: gentle-shake 0.3s ease-in-out;
    }

    .stDownloadButton button:hover {
        animation: shake 0.4s ease-in-out;
    }

    button[data-baseweb="tab"]:hover {
        animation: gentle-shake 0.2s ease-in-out;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white !important;
    }

    /* Stock cards */
    .stock-card {
        border: 1px solid #ddd;
        border-radius: 10px;
        padding: 15px;
        margin: 5px;
        background-color: white;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        transition: all 0.3s ease;
    }

    .stock-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.15);
    }

    .gradient-text {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        font-weight: bold;
    }

    /* Tabs */
    button[data-baseweb="tab"] {
        font-size: 18px !important;
        font-weight: 700 !important;
        font-family: 'Times New Roman', sans-serif !important;
        padding: 10px 20px !important;
    }

    /* Legend box */
    .legend-box {
        margin: 10px 0 20px 0;
        padding: 15px;
        background-color: #f8f9fa;
        border-radius: 10px;
        border: 1px solid #dee2e6;
    }

    .legend-title {
        margin: 0 0 10px 0;
        font-size: 16px;
        font-weight: bold;
    }

    .legend-item {
        display: flex;
        align-items: center;
        gap: 8px;
    }

    .legend-color {
        width: 20px;
        height: 20px;
        border-radius: 3px;
    }

    /* ============================================ */
    /* MOBILE RESPONSIVE STYLES - ENHANCED */
    /* ============================================ */
    @media only screen and (max-width: 768px) {
        /* Main container padding */
        .stApp {
            padding: 0.5rem !important;
        }

        /* Metric cards - full width on mobile */
        div[data-testid="stMetric"] {
            margin-bottom: 10px !important;
            padding: 10px !important;
        }

        /* Headers */
        h1 {
            font-size: 22px !important;
            text-align: center !important;
        }

        h3 {
            font-size: 18px !important;
        }

        h4 {
            font-size: 16px !important;
        }

        /* Tabs - scrollable on mobile */
        button[data-baseweb="tab"] {
            font-size: 12px !important;
            padding: 6px 10px !important;
            white-space: nowrap !important;
        }

        /* Tab container - horizontal scroll */
        [data-testid="stTabs"] {
            overflow-x: auto !important;
            white-space: nowrap !important;
            flex-wrap: nowrap !important;
        }

        [data-testid="stTabs"] button {
            flex: 0 0 auto !important;
        }

        /* Stock cards */
        .stock-card {
            padding: 8px !important;
            margin: 5px 0 !important;
        }

        .stock-card h4 {
            font-size: 14px !important;
        }

        .stock-card p {
            font-size: 12px !important;
            margin: 5px 0 !important;
        }

        /* Legend box */
        .legend-box {
            padding: 8px !important;
        }

        .legend-item {
            font-size: 10px !important;
        }

        .legend-color {
            width: 14px !important;
            height: 14px !important;
        }

        /* Dataframes - horizontal scroll */
        .stDataFrame {
            overflow-x: auto !important;
        }

        div[data-testid="stDataFrame"] {
            overflow-x: auto !important;
        }

        /* Metric values */
        div[data-testid="stMetricValue"] {
            font-size: 20px !important;
        }

        div[data-testid="stMetricLabel"] {
            font-size: 12px !important;
        }

        /* Buttons */
        .stButton button {
            font-size: 12px !important;
            padding: 6px 12px !important;
        }

        /* Sidebar */
        [data-testid="stSidebar"] {
            width: 280px !important;
        }

        [data-testid="stSidebar"] .stMarkdown {
            font-size: 12px !important;
        }

        /* Columns - stack on mobile */
        .row-widget.stHorizontal {
            flex-wrap: wrap !important;
        }

        /* Expander */
        details {
            font-size: 12px !important;
        }

        /* Download button caption */
        .stCaption {
            font-size: 10px !important;
        }

        /* Chart containers */
        .stPlotlyChart {
            width: 100% !important;
            overflow-x: auto !important;
        }

        /* Data editor */
        .stDataEditor {
            font-size: 11px !important;
        }

        /* Input fields */
        .stTextInput input {
            font-size: 14px !important;
            padding: 8px !important;
        }

        /* Select boxes */
        .stSelectbox div {
            font-size: 14px !important;
        }

        /* Info/Warning/Success messages */
        .stAlert {
            font-size: 12px !important;
            padding: 8px !important;
        }

        /* Metric delta */
        [data-testid="stMetricDelta"] {
            font-size: 10px !important;
        }
    }

    /* Extra small devices (phones below 480px) */
    @media only screen and (max-width: 480px) {
        h1 {
            font-size: 18px !important;
        }

        button[data-baseweb="tab"] {
            font-size: 10px !important;
            padding: 4px 8px !important;
        }

        div[data-testid="stMetricValue"] {
            font-size: 16px !important;
        }

        .stock-card h4 {
            font-size: 12px !important;
        }

        .stock-card p {
            font-size: 10px !important;
        }
    }

    /* Tablets (landscape) */
    @media only screen and (min-width: 769px) and (max-width: 1024px) {
        button[data-baseweb="tab"] {
            font-size: 14px !important;
            padding: 8px 14px !important;
        }

        h1 {
            font-size: 28px !important;
        }

        div[data-testid="stMetricValue"] {
            font-size: 22px !important;
        }
    }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------
# Initialize session state
# ---------------------------------------------------
if 'data_timestamp' not in st.session_state:
    st.session_state.data_timestamp = datetime.now()
if 'auto_refresh' not in st.session_state:
    st.session_state.auto_refresh = False
if 'recommendations' not in st.session_state:
    st.session_state.recommendations = {}
if 'heatmap_page' not in st.session_state:
    st.session_state.heatmap_page = 1
if 'google_sheets_data' not in st.session_state:
    st.session_state.google_sheets_data = None
if 'branch_amc_data' not in st.session_state:
    st.session_state.branch_amc_data = None
if 'supabase_client' not in st.session_state:
    st.session_state.supabase_client = init_supabase()
if 'branch_data' not in st.session_state:
    st.session_state.branch_data = None
if 'search_query' not in st.session_state:
    st.session_state.search_query = ""
if 'last_sheet_name' not in st.session_state:
    st.session_state.last_sheet_name = ""
if 'saved_recommendations' not in st.session_state:
    st.session_state.saved_recommendations = {}
if 'view_mode' not in st.session_state:
    st.session_state.view_mode = "table"
if 'risk_type_filter' not in st.session_state:
    st.session_state.risk_type_filter = "All"
if 'subcategory_filter' not in st.session_state:
    st.session_state.subcategory_filter = "All"
if 'previous_nsoh_data' not in st.session_state:
    st.session_state.previous_nsoh_data = None
if 'nsoh_changes' not in st.session_state:
    st.session_state.nsoh_changes = None
if 'raw_previous_data' not in st.session_state:
    st.session_state.raw_previous_data = None
if 'material_views' not in st.session_state:
    st.session_state.material_views = {}
if 'user_activity' not in st.session_state:
    st.session_state.user_activity = []
if 'notifications' not in st.session_state:
    st.session_state.notifications = []
if 'dos_tracking' not in st.session_state:
    st.session_state.dos_tracking = {}  # Stores current DOS days for each material

# Check connection periodically
if st.session_state.supabase_client and not check_supabase_connection():
    st.warning("⚠️ Supabase connection lost. Attempting to reconnect...")
    st.session_state.supabase_client = init_supabase()

# ---------------------------------------------------
# Program Hierarchy Configuration
# ---------------------------------------------------
PROGRAM_HIERARCHY = {
    "OI and Hepatitis": {
        "subcategories": ["AHD", "Hepatitis", "OI", "STI"],
        "is_parent": True
    },
    "TB": {
        "subcategories": ["Drug Susceptible -TB Medicine (DS-TB)", "Drug Resisitance -TB Medicine (DR-TB)", "Leprosy Medicines", "Nutrition"],
        "is_parent": True
    }
}

# ---------------------------------------------------
# Function to assign subcategories based on sequential order
# ---------------------------------------------------
def assign_subcategories_to_materials(df, subcategory_list):
    """Assign subcategory to each material based on sequential order in Material Description column"""
    subcategory_mapping = {}
    current_subcategory = None

    for idx in range(len(df)):
        material_desc = str(df.iloc[idx]['Material Description']).strip()

        # Check if this row is a subcategory header
        is_subcategory = False
        for subcat in subcategory_list:
            if material_desc == subcat:
                current_subcategory = subcat
                is_subcategory = True
                break

        # If not a subcategory header and we have a current subcategory, assign it
        if not is_subcategory and current_subcategory is not None:
            subcategory_mapping[material_desc] = current_subcategory

    return subcategory_mapping

def is_subcategory_header(material_desc, subcategory_list):
    """Check if a material description is actually a subcategory header"""
    return material_desc.strip() in subcategory_list

def filter_out_subcategory_headers(df, subcategory_list):
    """Remove subcategory header rows from the dataframe"""
    if subcategory_list is None or not subcategory_list:
        return df

    # Create a mask to filter out rows where Material Description is a subcategory header
    mask = ~df['Material Description'].astype(str).str.strip().isin(subcategory_list)
    return df[mask].copy()

# ---------------------------------------------------
# Expiry Risk Calculation Function
# ---------------------------------------------------
def parse_multiple_expiry_batches(expiry_str, amc):
    """Parse multiple batches and determine expiry risk"""
    try:
        if pd.isna(expiry_str) or expiry_str == "" or expiry_str is None:
            return False, ""

        expiry_str = str(expiry_str)
        pattern = r'(\d[\d,]*)\s*\(([A-Za-z]+)-(\d{4})\)'
        matches = re.findall(pattern, expiry_str)

        if not matches:
            return False, ""

        batches = []
        month_map = {'Jan':1, 'Feb':2, 'Mar':3, 'Apr':4, 'May':5, 'Jun':6,
                    'Jul':7, 'Aug':8, 'Sep':9, 'Oct':10, 'Nov':11, 'Dec':12}

        for quantity_str, month, year in matches:
            quantity = float(quantity_str.replace(',', ''))
            month_num = month_map.get(month[:3], 1)
            expiry_date = datetime(int(year), month_num, 1)
            batches.append({'quantity': quantity, 'expiry_date': expiry_date})

        batches.sort(key=lambda x: x['expiry_date'])

        if pd.isna(amc) or amc <= 0:
            return False, ""

        cumulative_stock = 0
        has_risk = False
        risk_details = []

        for batch in batches:
            months_until_expiry = max(0, (batch['expiry_date'].year - datetime.now().year) * 12 + 
                                      (batch['expiry_date'].month - datetime.now().month))
            cumulative_stock += batch['quantity']
            stock_needed = amc * months_until_expiry

            if cumulative_stock > stock_needed:
                excess = cumulative_stock - stock_needed
                batch_risk = min(batch['quantity'], excess)
                if batch_risk > 0:
                    has_risk = True
                    risk_details.append(f"{batch_risk:,.0f} units expiring {batch['expiry_date'].strftime('%b-%Y')}")

        return has_risk, "; ".join(risk_details) if risk_details else ""
    except Exception as e:
        return False, f"Error: {e}"

# Database Connection Functions
# ---------------------------------------------------
def load_national_data():
    """Load ALL national_data from Supabase with pagination"""
    try:
        if st.session_state.supabase_client is None:
            st.error("Supabase client not initialized")
            return pd.DataFrame()

        # Pagination to get ALL rows
        all_data = []
        page = 0
        page_size = 1000

        while True:
            response = st.session_state.supabase_client.table("health_data") \
                .select("*") \
                .range(page * page_size, (page + 1) * page_size - 1) \
                .execute()

            if not response.data:
                break

            all_data.extend(response.data)

            if len(response.data) < page_size:
                break

            page += 1

        if not all_data:
            st.warning("No data found in Supabase")
            return pd.DataFrame()

        df = pd.DataFrame(all_data)

        if 'id' in df.columns:
            df = df.drop(columns=['id'])

        column_mapping = {
            'material_description': 'Material Description',
            'adama_branch': 'Adama Branch',
            'addis_ababa_branch_1': 'Addis Ababa Branch 1',
            'addis_ababa_branch_2': 'Addis Ababa Branch 2',
            'arba_minch_branch': 'Arba Minch Branch',
            'assosa_branch': 'Assosa Branch',
            'bahir_dar_branch': 'Bahir Dar Branch',
            'dessie_branch': 'Dessie Branch',
            'dire_dawa_branch': 'Dire Dawa Branch',
            'gambela_branch': 'Gambela Branch',
            'gondar_branch': 'Gondar Branch',
            'hawassa_branch': 'Hawassa Branch',
            'jigjiga_branch': 'Jigjiga Branch',
            'jimma_branch': 'Jimma Branch',
            'kebridahar_branch': 'Kebridahar Branch',
            'mekele_branch': 'Mekele Branch',
            'negele_borena_branch': 'Negele Borena Branch',
            'nekemte_branch': 'Nekemte Branch',
            'semera_branch': 'Semera Branch',
            'shire_branch': 'Shire Branch',
            'head_office': 'Head Office',
            'hubs': 'Hubs',
            'nsoh': 'NSOH',
            'expiry': 'Expiry'
        }

        existing_mapping = {k: v for k, v in column_mapping.items() if k in df.columns}
        df = df.rename(columns=existing_mapping)

        df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
        df.columns = df.columns.str.strip()

        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].fillna("")

        return df

    except Exception as e:
        st.error(f"Error loading data from Supabase: {e}")
        return pd.DataFrame()
@st.cache_data(ttl=300, show_spinner=False)
def load_new_deliveries():
    """Load new_deliveries data from Supabase"""
    try:
        if st.session_state.supabase_client is None:
            return pd.DataFrame()
        response = st.session_state.supabase_client.table("new_deliveries").select("*").execute()
        if response.data:
            df = pd.DataFrame(response.data)
            if 'id' in df.columns:
                df = df.drop(columns=['id'])
            # Rename columns for display
            column_rename = {
                'material_description': 'Material Description',
                'posting_date': 'Posting Date',
                'purchase_order': 'PO Number',
                'quantity': 'Quantity'
            }
            existing_rename = {k: v for k, v in column_rename.items() if k in df.columns}
            df = df.rename(columns=existing_rename)
            return df
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading New_deliveries data: {e}")
        return pd.DataFrame()
@st.cache_data(ttl=300, show_spinner=False)
def load_branch_amc_from_google_sheets(sheet_id):
    """Load branch AMC data from Google Sheets"""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

    for attempt in range(1):
        try:
            session = requests.Session()
            session.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})
            session.trust_env = False

            response = session.get(url, timeout=45)
            response.raise_for_status()

            content = response.content
            df = pd.read_excel(BytesIO(content), header=0)

            df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
            df.columns = df.columns.str.strip()

            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].fillna("")

            st.session_state.branch_amc_data = df
            return df
        except Exception as e:
            if st.session_state.branch_amc_data is not None:
                st.warning(f"Using cached branch AMC data. New data unavailable: {str(e)}")
                return st.session_state.branch_amc_data
            st.warning(f"Could not load branch AMC data: {str(e)}")
            return pd.DataFrame(columns=['Material Description'])
    return pd.DataFrame()

@st.cache_data(ttl=300, show_spinner=False)
def load_google_sheets(sheet_id):
    """Load Google Sheets data (AMC and pipeline data)"""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

    for attempt in range(3):
        try:
            session = requests.Session()
            session.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})
            session.trust_env = False

            response = session.get(url, timeout=45)
            response.raise_for_status()

            content = response.content
            sheets = pd.read_excel(BytesIO(content), sheet_name=None, header=2)

            cleaned_sheets = {}
            for name, df in sheets.items():
                if df.empty:
                    cleaned_sheets[name] = df
                    continue

                df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
                df.columns = df.columns.str.strip()

                for col in df.columns:
                    try:
                        if col in df.columns and len(df) > 0:
                            if pd.api.types.is_object_dtype(df[col]):
                                df[col] = df[col].fillna("")
                            elif df[col].dtype == 'object':
                                df[col] = df[col].fillna("")
                    except Exception:
                        continue

                cleaned_sheets[name] = df

            st.session_state.google_sheets_data = cleaned_sheets
            return cleaned_sheets
        except Exception as e:
            if attempt < 2:
                time.sleep(2)
            else:
                if st.session_state.google_sheets_data:
                    st.warning(f"Using cached Google Sheets data. New data unavailable: {str(e)}")
                    return st.session_state.google_sheets_data
                st.error(f"Error loading Google Sheets: {str(e)}")
                return {}
    return {}

def validate_upload_data(df):
    """Validate data before uploading to Supabase"""
    required_columns = ['Material Description', 'NSOH']
    missing_columns = [col for col in required_columns if col not in df.columns]

    if missing_columns:
        st.error(f"Missing required columns: {', '.join(missing_columns)}")
        return False

    if df['Material Description'].duplicated().any():
        st.warning("Duplicate Material Descriptions found. Only first occurrence will be kept.")
        df = df.drop_duplicates(subset=['Material Description'], keep='first')

    return True

# ---------------------------------------------------
# Admin Functions for Supabase
# ---------------------------------------------------
def upload_to_supabase(df, table_name="health_data"):
    """Upload DataFrame to Supabase"""
    try:
        if st.session_state.supabase_client is None:
            st.error("Supabase client not initialized")
            return False

        reverse_mapping = {
            'Material Description': 'material_description',
            'Adama Branch': 'adama_branch',
            'Addis Ababa Branch 1': 'addis_ababa_branch_1',
            'Addis Ababa Branch 2': 'addis_ababa_branch_2',
            'Arba Minch Branch': 'arba_minch_branch',
            'Assosa Branch': 'assosa_branch',
            'Bahir Dar Branch': 'bahir_dar_branch',
            'Dessie Branch': 'dessie_branch',
            'Dire Dawa Branch': 'dire_dawa_branch',
            'Gambela Branch': 'gambela_branch',
            'Gondar Branch': 'gondar_branch',
            'Hawassa Branch': 'hawassa_branch',
            'Jigjiga Branch': 'jigjiga_branch',
            'Jimma Branch': 'jimma_branch',
            'Kebridahar Branch': 'kebridahar_branch',
            'Mekele Branch': 'mekele_branch',
            'Negele Borena Branch': 'negele_borena_branch',
            'Nekemte Branch': 'nekemte_branch',
            'Semera Branch': 'semera_branch',
            'Shire Branch': 'shire_branch',
            'Head Office': 'head_office',
            'Hubs': 'hubs',
            'NSOH': 'nsoh',
            'Expiry': 'expiry'
        }

        upload_columns = [col for col in reverse_mapping.keys() if col in df.columns]
        df_upload = df[upload_columns].copy()
        df_upload = df_upload.rename(columns=reverse_mapping)
        df_upload = df_upload.replace({np.nan: None})
        data = df_upload.to_dict(orient="records")

        response = st.session_state.supabase_client.table(table_name).insert(data).execute()

        st.success(f"Successfully uploaded {len(data)} records to Supabase!")
        return True
    except Exception as e:
        st.error(f"Error uploading to Supabase: {e}")
        return False

def clear_supabase_table(table_name="health_data"):
    """Clear all data from Supabase table"""
    try:
        if st.session_state.supabase_client is None:
            st.error("Supabase client not initialized")
            return False

        response = st.session_state.supabase_client.table(table_name).delete().neq("id", 0).execute()

        st.success(f"Successfully cleared table {table_name}!")
        return True
    except Exception as e:
        st.error(f"Error clearing table: {e}")
        return False

def get_table_info():
    """Get information about the Supabase table"""
    try:
        if st.session_state.supabase_client is None:
            return None

        response = st.session_state.supabase_client.table("health_data").select("*").execute()

        if response.data:
            df = pd.DataFrame(response.data)
            return {"rows": len(df), "columns": len(df.columns), "column_names": list(df.columns)}
        return None
    except Exception as e:
        st.error(f"Error getting table info: {e}")
        return None

# ---------------------------------------------------
# Stock Change Tracking Function (Updated to use NMOS)
# ---------------------------------------------------
def calculate_stock_changes(current_df, previous_df):
    """Calculate stock quantity changes based on NSOH and NMOS - shows ALL materials"""

    if previous_df is None or previous_df.empty:
        return None

    # Get NSOH and NMOS from both dataframes
    current_data = current_df[['Material Description', 'NSOH', 'NMOS']].copy()
    previous_data = previous_df[['Material Description', 'NSOH', 'NMOS']].copy()

    # Convert to numeric and round NSOH to whole numbers
    for col in ['NSOH', 'NMOS']:
        current_data[col] = pd.to_numeric(current_data[col], errors='coerce').fillna(0)
        previous_data[col] = pd.to_numeric(previous_data[col], errors='coerce').fillna(0)

    # Round NSOH to whole numbers
    current_data['NSOH'] = current_data['NSOH'].round(0)
    previous_data['NSOH'] = previous_data['NSOH'].round(0)

    # Merge (use outer join to include all materials from both datasets)
    merged = current_data.merge(
        previous_data, 
        on='Material Description', 
        suffixes=('_now', '_previous'),
        how='outer'  # This ensures ALL materials are included
    )

    # Fill NaN values with 0 for materials that only exist in one dataset
    merged['NSOH_now'] = merged['NSOH_now'].fillna(0)
    merged['NSOH_previous'] = merged['NSOH_previous'].fillna(0)
    merged['NMOS_now'] = merged['NMOS_now'].fillna(0)
    merged['NMOS_previous'] = merged['NMOS_previous'].fillna(0)

    # Calculate changes
    merged['NSOH_Change'] = merged['NSOH_now'] - merged['NSOH_previous']
    merged['NMOS_Change'] = merged['NMOS_now'] - merged['NMOS_previous']

    # Calculate percentage change for NSOH
    merged['NSOH_Change_Pct'] = np.where(
        merged['NSOH_previous'] > 0,
        (merged['NSOH_Change'] / merged['NSOH_previous']) * 100,
        0
    )

    # Determine change type (includes NO CHANGE)
    def get_change_type(row):
        if row['NSOH_Change'] > 0:
            return "📦 STOCK ADDED"
        elif row['NSOH_Change'] < 0:
            return "📉 STOCK CONSUMED"
        else:
            return "➖ NO CHANGE"

    merged['Change Type'] = merged.apply(get_change_type, axis=1)

    # Format numbers
    merged['NSOH_previous'] = merged['NSOH_previous'].apply(lambda x: f"{int(x):,}")
    merged['NSOH_now'] = merged['NSOH_now'].apply(lambda x: f"{int(x):,}")
    merged['NSOH_Change'] = merged['NSOH_Change'].apply(lambda x: f"{'+' if x > 0 else ''}{int(x):,}" if x != 0 else "0")
    merged['NSOH_Change_Pct'] = merged['NSOH_Change_Pct'].apply(lambda x: f"{x:+.1f}%" if x != 0 else "0%")
    merged['NMOS_previous'] = merged['NMOS_previous'].apply(lambda x: f"{x:.2f}")
    merged['NMOS_now'] = merged['NMOS_now'].apply(lambda x: f"{x:.2f}")
    merged['NMOS_Change'] = merged['NMOS_Change'].apply(lambda x: f"{x:+.2f}" if x != 0 else "0.00")

    # Return single table with all change data (including NO CHANGE)
    result = merged[[
        'Material Description',
        'NSOH_previous', 'NSOH_now', 'NSOH_Change', 'NSOH_Change_Pct',
        'NMOS_previous', 'NMOS_now', 'NMOS_Change',
        'Change Type'
    ]].copy()

    # Sort by absolute NSOH change (largest first, NO CHANGE at bottom)
    result['_sort_key'] = result['NSOH_Change'].str.replace(',', '').str.replace('+', '').astype(int).abs()
    result = result.sort_values('_sort_key', ascending=False).drop('_sort_key', axis=1)

    return result

# ---------------------------------------------------
# Utility Functions
# ---------------------------------------------------
def format_number_with_commas(x):
    try:
        if pd.isna(x) or x == "" or x is None:
            return ""
        if isinstance(x, str):
            x = x.replace(',', '')
            try:
                x = float(x) if x else np.nan
            except:
                return x
        if pd.isna(x):
            return ""
        x = round(x)
        return f"{x:,.0f}"
    except:
        return str(x) if x else ""

def format_mos_with_decimals(x):
    try:
        if pd.isna(x) or x == "" or x is None:
            return ""
        if isinstance(x, str):
            try:
                x = float(x) if x else np.nan
            except:
                return x
        if pd.isna(x):
            return ""
        return f"{x:.2f}"
    except:
        return str(x) if x else ""

def categorize_stock(nmos):
    try:
        if pd.isna(nmos) or nmos == "" or nmos is None:
            return ""
        x = float(nmos) if not isinstance(nmos, (int, float)) else nmos
        if x < 1:
            return "Stock Out"
        elif x < 6:
            return "Understock"
        elif x <= 18:
            return "Normal Stock"
        else:
            return "Overstock"
    except:
        return ""

def calculate_coefficient_of_variation(values):
    try:
        values = pd.to_numeric(values, errors='coerce')
        values = values[values.notna() & (values > 0)]
        if len(values) > 1:
            mean = values.mean()
            std = values.std()
            if mean > 0:
                return (std / mean) * 100
        return np.nan
    except:
        return np.nan

def calculate_risk(row):
    """Calculate risk of stock out - ONLY for NMOS >= 1 (Stock Out is separate)"""
    try:
        nmos = row['NMOS'] if pd.notna(row['NMOS']) else np.nan
        git_mos = row['GIT_MOS'] if pd.notna(row['GIT_MOS']) else 0
        lc_mos = row['LC_MOS'] if pd.notna(row['LC_MOS']) else 0
        wb_mos = row['WB_MOS'] if pd.notna(row['WB_MOS']) else 0
        tmd_mos = row['TMD_MOS'] if pd.notna(row['TMD_MOS']) else 0

        if pd.isna(nmos) or nmos < 1:
            return ""

        if 1 <= nmos < 2:
            return "Risk of Stock out"

        if 2 <= nmos < 4 and git_mos == 0:
            return "Risk of Stock out"

        if 4 <= nmos < 6 and git_mos == 0 and lc_mos == 0 and wb_mos == 0:
            return "Risk of Stock out"

        if 6 <= nmos < 7 and git_mos == 0 and lc_mos == 0 and wb_mos == 0 and tmd_mos == 0:
            return "Risk of Stock out"

        return ""
    except:
        return ""
def calculate_dos(current_df):
    """Calculate Days Out of Stock - counts actual calendar days"""
    from datetime import date

    current_date = date.today()

    # Initialize if not exists
    if 'dos_tracking' not in st.session_state:
        st.session_state.dos_tracking = {}

    for idx, row in current_df.iterrows():
        material = row.get('Material Description')
        if pd.isna(material):
            continue

        nmos = row.get('NMOS', 1)
        try:
            nmos = float(nmos) if pd.notna(nmos) else 1
        except:
            nmos = 1

        is_out_of_stock = nmos < 1

        if material not in st.session_state.dos_tracking:
            # Initialize tracking for this material
            if is_out_of_stock:
                st.session_state.dos_tracking[material] = {
                    'start_date': current_date,
                    'days': 1
                }
            else:
                st.session_state.dos_tracking[material] = {
                    'start_date': None,
                    'days': 0
                }
        else:
            tracking = st.session_state.dos_tracking[material]

            if is_out_of_stock:
                if tracking['start_date'] is None:
                    # Just went out of stock today
                    tracking['start_date'] = current_date
                    tracking['days'] = 1
                else:
                    # Calculate days including today
                    tracking['days'] = (current_date - tracking['start_date']).days + 1
            else:
                # Back in stock - reset
                if tracking['days'] > 0:
                    # Record the episode for history (optional)
                    if 'dos_history' not in st.session_state:
                        st.session_state.dos_history = {}
                    if material not in st.session_state.dos_history:
                        st.session_state.dos_history[material] = []
                    st.session_state.dos_history[material].append({
                        'start': tracking['start_date'],
                        'end': current_date,
                        'days': tracking['days']
                    })

                # Reset
                tracking['start_date'] = None
                tracking['days'] = 0

    # Return days only
    return {material: st.session_state.dos_tracking[material]['days'] 
            for material in st.session_state.dos_tracking}
def get_stock_out_recommendation(row):
    """Generate recommendation for Risk of Stock out based on pipeline status with PO numbers"""
    try:
        git_mos = row.get('GIT_MOS', 0)
        lc_mos = row.get('LC_MOS', 0)
        wb_mos = row.get('WB_MOS', 0)
        tmd_mos = row.get('TMD_MOS', 0)
        git_po = row.get('GIT_PO', '')
        lc_po = row.get('LC_PO', '')
        wb_po = row.get('WB_PO', '')
        tmd_po = row.get('TMD_PO', '')

        git_mos = pd.to_numeric(git_mos, errors='coerce') if not isinstance(git_mos, (int, float)) else git_mos
        lc_mos = pd.to_numeric(lc_mos, errors='coerce') if not isinstance(lc_mos, (int, float)) else lc_mos
        wb_mos = pd.to_numeric(wb_mos, errors='coerce') if not isinstance(wb_mos, (int, float)) else wb_mos
        tmd_mos = pd.to_numeric(tmd_mos, errors='coerce') if not isinstance(tmd_mos, (int, float)) else tmd_mos

        git_mos = git_mos if pd.notna(git_mos) else 0
        lc_mos = lc_mos if pd.notna(lc_mos) else 0
        wb_mos = wb_mos if pd.notna(wb_mos) else 0
        tmd_mos = tmd_mos if pd.notna(tmd_mos) else 0

        if git_mos > 0:
            po_text = f" PO {git_po}" if git_po and str(git_po) != 'nan' and str(git_po) != '' else ""
            return f"🚚 Expedite shipment{po_text} - goods in transit"
        elif git_mos == 0 and lc_mos > 0:
            po_text = f" PO {lc_po}" if lc_po and str(lc_po) != 'nan' and str(lc_po) != '' else ""
            return f"📄 Expedite L/C opening process{po_text}"
        elif git_mos == 0 and lc_mos == 0 and wb_mos > 0:
            po_text = f" PO {wb_po}" if wb_po and str(wb_po) != 'nan' and str(wb_po) != '' else ""
            return f"💰 Expedite budget transfer{po_text}"
        elif git_mos == 0 and lc_mos == 0 and wb_mos == 0 and tmd_mos > 0:
            po_text = f" PO {tmd_po}" if tmd_po and str(tmd_po) != 'nan' and str(tmd_po) != '' else ""
            return f"📋 Expedite tender process{po_text}"
        else:
            return "🔄 Initiate additional quantity - no pipeline stock"
    except Exception as e:
        return f"⚠️ Review supply chain status"

def get_expiry_risk_recommendation(row, cv_category=None):
    """Generate recommendation for Expiry Risk based on CV and distribution, including expiry quantity"""
    try:
        if cv_category is None:
            cv_category = row.get('CV Category', 'Unknown')

        hubs_pct = row.get('Hubs%', 0)
        ho_pct = row.get('Head Office%', 0)
        expiry_details = row.get('Expiry Risk Details', '')

        hubs_pct = pd.to_numeric(hubs_pct, errors='coerce') if not isinstance(hubs_pct, (int, float)) else hubs_pct
        ho_pct = pd.to_numeric(ho_pct, errors='coerce') if not isinstance(ho_pct, (int, float)) else ho_pct
        hubs_pct = hubs_pct if pd.notna(hubs_pct) else 0
        ho_pct = ho_pct if pd.notna(ho_pct) else 0

        expiry_info = f" [Expiry: {expiry_details}]" if expiry_details and expiry_details != "" else ""

        if cv_category == 'High variation':
            return f"🔄 Redistribution required - high variation across hubs{expiry_info}"
        elif hubs_pct < ho_pct:
            return f"📦 Push stock to hubs - hubs have lower stock than head office{expiry_info}"
        else:
            return f"🌍 Donate to other countries - excess stock for donation{expiry_info}"

    except Exception as e:
        return f"⚠️ Review expiry risk - check batch details"

# ---------------------------------------------------
# Load ALL data
# ---------------------------------------------------
amc_pipeline_sheet_id = "14VvZ7IyOmpM4SZrY5_ArHDgLkeFN4inW"
branch_amc_sheet_id = "12Z5xqX32QIzjoN6tNvGbjutMheXx5US1"

df_external = load_national_data()
df_new_deliveries = load_new_deliveries()
branch_amc_data = load_branch_amc_from_google_sheets(branch_amc_sheet_id)
google_sheets = load_google_sheets(amc_pipeline_sheet_id)

if df_external.empty:
    st.error("No data in Supabase. Please upload data through admin panel.")
    st.stop()

# ---------------------------------------------------
# User Info in Sidebar
# ---------------------------------------------------
with st.sidebar:
    st.title(f"Welcome, {st.session_state['user']['full_name']}!")
    st.caption(f"Role: {st.session_state['user']['role'].title()}")

    if st.session_state['user']['role'] == 'admin' and st.session_state.supabase_client:
        st.success("✅ Connected to Supabase")

    if st.session_state.notifications:
        st.markdown(f"### 🔔 Notifications ({len(st.session_state.notifications)})")
        for notif in st.session_state.notifications[:3]:
            st.warning(notif)

# ---------------------------------------------------
# Program Selection
# ---------------------------------------------------
if google_sheets:
    program_list = ["All"] + list(google_sheets.keys())
else:
    program_list = ["All"]

sheet_name = st.sidebar.selectbox("Program", program_list, index=0, key="program_selector")

subcategory_options = ["All"]
if sheet_name in PROGRAM_HIERARCHY and PROGRAM_HIERARCHY[sheet_name]["is_parent"]:
    subcategory_options = ["All"] + PROGRAM_HIERARCHY[sheet_name]["subcategories"]
    subcategory_filter = st.sidebar.selectbox("Subcategory", subcategory_options, key="subcategory_filter")
else:
    subcategory_filter = "All"
    st.session_state.subcategory_filter = "All"

if sheet_name != st.session_state.last_sheet_name:
    st.session_state.heatmap_page = 1
    st.session_state.last_sheet_name = sheet_name

if sheet_name == "All" and google_sheets:
    all_dfs = []
    for name, df_program in google_sheets.items():
        if df_program.empty:
            continue
        df_copy = df_program.copy()
        df_copy.columns = df_copy.columns.astype(str)
        if df_copy.columns.duplicated().any():
            df_copy = df_copy.loc[:, ~df_copy.columns.duplicated()]
        all_dfs.append(df_copy)

    if all_dfs:
        try:
            df_google = pd.concat(all_dfs, ignore_index=True, sort=False)
        except Exception as concat_error:
            st.error(f"Error combining sheets: {concat_error}")
            df_google = pd.DataFrame()
    else:
        df_google = pd.DataFrame()
elif google_sheets and sheet_name in google_sheets:
    df_google = google_sheets[sheet_name].copy()
    df_google.columns = df_google.columns.astype(str)
    if df_google.columns.duplicated().any():
        df_google = df_google.loc[:, ~df_google.columns.duplicated()]
else:
    df_google = pd.DataFrame()

if not df_google.empty:
    required_cols = [
        'Material Description', 'AMC',
        'GIT_PO', 'GIT_Qty', 'GIT_MOS',
        'LC_PO', 'LC_Qty', 'LC_MOS',
        'WB_PO', 'WB_Qty', 'WB_MOS',
        'TMD_PO', 'TMD_Qty', 'TMD_MOS', "Status"
    ]
    available_cols = [c for c in required_cols if c in df_google.columns]
    df_google = df_google[available_cols]
else:
    df_google = pd.DataFrame()

if not df_google.empty and not df_external.empty:
    if 'Material Description' in df_external.columns and 'Material Description' in df_google.columns:
        df_external = df_external.drop_duplicates(subset=['Material Description'], keep='first')
        df_google = df_google.drop_duplicates(subset=['Material Description'], keep='first')
        df = df_external.merge(df_google, on="Material Description", how="right")
        df = df.drop_duplicates(subset=['Material Description'], keep='first')
    else:
        st.error("Material Description column missing for merge")
        df = df_external.copy()
else:
    df = df_external.copy()

if not df.empty:
    if 'S/N' in df.columns:
        df = df.drop(columns=['S/N'])

    text_columns = ['Status', 'Expiry', 'GIT_PO', 'LC_PO', 'WB_PO', 'TMD_PO']
    numeric_columns = [col for col in df.columns if col not in text_columns + ['Material Description']]

    for col in numeric_columns:
        if col in df.columns:
            try:
                df[col] = pd.to_numeric(df[col], errors='coerce')
            except Exception:
                continue

    for col in text_columns:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str)

    if 'NSOH' in df.columns and 'AMC' in df.columns:
        nsoh = pd.to_numeric(df['NSOH'], errors='coerce')
        amc = pd.to_numeric(df['AMC'], errors='coerce')
        amc = amc.replace(0, np.nan)
        nmos = nsoh / amc
        nmos = nmos.replace([np.inf, -np.inf], np.nan)
        df['NMOS'] = nmos.round(2)

    # Calculate DOS
    dos_dict = calculate_dos(df)
    df['DOS'] = df['Material Description'].map(dos_dict).fillna(0).astype(int)

    mos_cols = ['NMOS', 'GIT_MOS', 'LC_MOS', 'WB_MOS', 'TMD_MOS']
    available_mos = []
    for col in mos_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
            available_mos.append(col)

    if available_mos:
        df['TMOS'] = df[available_mos].sum(axis=1)
    else:
        df['TMOS'] = np.nan

    if 'NMOS' in df.columns:
        df['Stock Status'] = df['NMOS'].apply(categorize_stock)
    else:
        df['Stock Status'] = ""

    if 'Hubs' in df.columns and 'Head Office' in df.columns and 'NSOH' in df.columns:
        hubs_vals = pd.to_numeric(df['Hubs'], errors='coerce').fillna(0)
        ho_vals = pd.to_numeric(df['Head Office'], errors='coerce').fillna(0)
        nsoh_vals = pd.to_numeric(df['NSOH'], errors='coerce')
        valid_mask = nsoh_vals.notna() & (nsoh_vals > 0)
        df['Hubs%'] = np.where(valid_mask, (hubs_vals / nsoh_vals * 100).round(1), np.nan)
        df['Head Office%'] = np.where(valid_mask, (ho_vals / nsoh_vals * 100).round(1), np.nan)
        df['Avail Gap'] = np.where(valid_mask, (df['Hubs%'] - df['Head Office%']).round(1), np.nan)
    else:
        df['Hubs%'] = np.nan
        df['Head Office%'] = np.nan
        df['Avail Gap'] = np.nan

    if not branch_amc_data.empty and 'Material Description' in branch_amc_data.columns:
        branch_cols = [col for col in df.columns if 'Branch' in col or col == 'Material Description']
        if len(branch_cols) > 1:
            stock_data = df[branch_cols].copy()
            merged_cv = pd.merge(stock_data, branch_amc_data, on='Material Description', how='inner', suffixes=('_stock', '_amc'))

            if not merged_cv.empty:
                branch_cols_list = [col for col in stock_data.columns if col != 'Material Description']
                amc_cols = [col for col in branch_amc_data.columns if col != 'Material Description']
                cv_calc_data = {'Material Description': merged_cv['Material Description']}

                for i in range(min(len(branch_cols_list), len(amc_cols))):
                    stock_col = branch_cols_list[i]
                    amc_col = amc_cols[i]
                    stock_vals = pd.to_numeric(merged_cv[f"{stock_col}_stock"], errors='coerce')
                    amc_vals = pd.to_numeric(merged_cv[f"{amc_col}_amc"], errors='coerce')
                    with np.errstate(divide='ignore', invalid='ignore'):
                        mos_vals = np.where(amc_vals > 0, stock_vals / amc_vals, np.nan)
                    cv_calc_data[stock_col] = mos_vals

                cv_df = pd.DataFrame(cv_calc_data)
                mos_cols_for_cv = [col for col in cv_df.columns if col != 'Material Description']
                cv_df['CV (%)'] = cv_df[mos_cols_for_cv].apply(lambda row: calculate_coefficient_of_variation(row), axis=1)
                cv_df['CV (%)'] = cv_df['CV (%)'].round(1)

                def categorize_cv(cv_value):
                    if pd.isna(cv_value):
                        return "Unknown"
                    elif cv_value < 50:
                        return "Low variation"
                    elif cv_value <= 100:
                        return "Moderate variation"
                    else:
                        return "High variation"

                cv_df['CV Category'] = cv_df['CV (%)'].apply(categorize_cv)
                df = df.merge(cv_df[['Material Description', 'CV Category']], on='Material Description', how='left')
            else:
                df['CV Category'] = "Unknown"
        else:
            df['CV Category'] = "Unknown"
    else:
        df['CV Category'] = "Unknown"

        # Handle subcategory filtering BEFORE other operations
    if sheet_name in PROGRAM_HIERARCHY:
        subcategory_list = PROGRAM_HIERARCHY[sheet_name]["subcategories"]

        # DO NOT filter out subcategory headers - keep them visible
        # df = filter_out_subcategory_headers(df, subcategory_list)

        # Assign subcategories to remaining materials
        subcategory_mapping = assign_subcategories_to_materials(df, subcategory_list)
        df['Assigned Subcategory'] = df['Material Description'].map(subcategory_mapping)

        # Apply subcategory filter if selected
        if subcategory_filter != "All":
            df = df[df['Assigned Subcategory'] == subcategory_filter]
    else:
        df['Assigned Subcategory'] = None

    if 'NMOS' in df.columns:
        df['Risk of Stock'] = df.apply(calculate_risk, axis=1)
    else:
        df['Risk of Stock'] = ""

    expiry_data = df.apply(lambda row: parse_multiple_expiry_batches(row.get('Expiry', ''), row.get('AMC', np.nan)), axis=1)
    df['Has Expiry Risk'] = expiry_data.apply(lambda x: x[0])
    df['Expiry Risk Details'] = expiry_data.apply(lambda x: x[1])

    risk_types = []
    for idx, row in df.iterrows():
        risk_of_stock = row.get('Risk of Stock', '') == 'Risk of Stock out'
        expiry_risk = row.get('Has Expiry Risk', False)

        if risk_of_stock and expiry_risk:
            risk_types.append("Critical Risk")
        elif risk_of_stock:
            risk_types.append("Risk of Stock out")
        elif expiry_risk:
            risk_types.append("Expiry Risk")
        else:
            risk_types.append("")
    df['Risk Type'] = risk_types

    # ---------------------------------------------------
    # Track stock changes AFTER NMOS is calculated
    # ---------------------------------------------------
    if st.session_state.raw_previous_data is not None:
        stock_changes = calculate_stock_changes(df, st.session_state.raw_previous_data)
        if stock_changes is not None:
            st.session_state.nsoh_changes = stock_changes
        else:
            st.session_state.nsoh_changes = None
    else:
        st.session_state.nsoh_changes = None

    # Store current data for next comparison
    st.session_state.raw_previous_data = df.copy()

    display_df = df.copy()
    text_columns_to_preserve = ['Material Description', 'Stock Status', 'Risk Type', 'Status', 'Expiry', 'GIT_PO', 'LC_PO', 'WB_PO', 'TMD_PO', 'CV Category','DOS']

    if 'Hubs%' in display_df.columns:
        text_columns_to_preserve.append('Hubs%')
    if 'Head Office%' in display_df.columns:
        text_columns_to_preserve.append('Head Office%')
    if 'Avail Gap' in display_df.columns:
        text_columns_to_preserve.append('Avail Gap')

    for col in display_df.columns:
        if col not in text_columns_to_preserve:
            if col in ['NMOS', 'GIT_MOS', 'LC_MOS', 'WB_MOS', 'TMD_MOS', 'TMOS']:
                display_df[col] = display_df[col].apply(format_mos_with_decimals)
            elif col in ['Hubs%', 'Head Office%', 'Avail Gap']:
                display_df[col] = display_df[col].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
            else:
                display_df[col] = display_df[col].apply(format_number_with_commas)

    # ========== MATERIAL FILTER SECTION ==========
if 'Material Description' in df.columns:
    # FIXED: Handle NaN values in Material Description for sorting
    unique_materials = df['Material Description'].dropna().astype(str).unique()

    # Filter out subcategory headers from material list (dropdown only)
    if sheet_name in PROGRAM_HIERARCHY:
        subcategory_list = PROGRAM_HIERARCHY[sheet_name]["subcategories"]
        unique_materials = [m for m in unique_materials if m not in subcategory_list]

    materials = ["All"] + sorted(unique_materials)

    statuses = ["All"] + sorted([s for s in df['Stock Status'].unique() if s != "" and pd.notna(s)]) if 'Stock Status' in df.columns else ["All"]

    risk_type_options = ["All", "Risk of Stock out", "Expiry Risk", "Critical Risk"]
    risk_type_filter = st.sidebar.selectbox("Risk Type", risk_type_options, index=risk_type_options.index(st.session_state.risk_type_filter) if st.session_state.risk_type_filter in risk_type_options else 0)
    st.session_state.risk_type_filter = risk_type_filter

    material_filter = st.sidebar.selectbox("Material Description", materials)
    status_filter = st.sidebar.selectbox("Stock Status", statuses)

    df_filtered = df.copy()
    display_df_filtered = display_df.copy()

    # ========== REMOVE SUBCATEGORY HEADERS FROM DATA ==========
    if sheet_name in PROGRAM_HIERARCHY:
        subcategory_list = PROGRAM_HIERARCHY[sheet_name]["subcategories"]
        mask = ~df_filtered['Material Description'].astype(str).str.strip().isin(subcategory_list)
        df_filtered = df_filtered[mask].copy()
        display_df_filtered = display_df_filtered[mask].copy()
    # ========== END REMOVE SUBCATEGORY HEADERS ==========

    if material_filter != "All":
        df_filtered = df_filtered[df_filtered['Material Description'] == material_filter]
        display_df_filtered = display_df_filtered[display_df_filtered['Material Description'] == material_filter]
        # Track material view for popular materials analytics
        if material_filter not in st.session_state.material_views:
            st.session_state.material_views[material_filter] = 0
        st.session_state.material_views[material_filter] += 1

        # Track user activity
        if 'user' in st.session_state:
            st.session_state.user_activity.append({
                'user': st.session_state['user']['email'],
                'role': st.session_state['user']['role'],
                'action': 'view_material',
                'material': material_filter,
                'timestamp': datetime.now().isoformat()
            })

    if status_filter != "All" and 'Stock Status' in df_filtered.columns:
        df_filtered = df_filtered[df_filtered['Stock Status'] == status_filter]
        display_df_filtered = display_df_filtered[display_df_filtered['Stock Status'] == status_filter]

    if risk_type_filter == "Risk of Stock out":
        df_filtered = df_filtered[df_filtered['Risk Type'] == "Risk of Stock out"]
        display_df_filtered = display_df_filtered[display_df_filtered['Risk Type'] == "Risk of Stock out"]
    elif risk_type_filter == "Expiry Risk":
        df_filtered = df_filtered[df_filtered['Risk Type'] == "Expiry Risk"]
        display_df_filtered = display_df_filtered[display_df_filtered['Risk Type'] == "Expiry Risk"]
    elif risk_type_filter == "Critical Risk":
        df_filtered = df_filtered[df_filtered['Risk Type'] == "Critical Risk"]
        display_df_filtered = display_df_filtered[display_df_filtered['Risk Type'] == "Critical Risk"]
else:
    st.error("Material Description column not found in the data")
    df_filtered = pd.DataFrame()
    display_df_filtered = pd.DataFrame()

# ---------------------------------------------------
# Navigation
# ---------------------------------------------------
st.sidebar.divider()

if st.session_state['user']['role'] == 'admin':
    page = st.sidebar.radio("Navigation", ["Dashboard", "Advanced Analytics", "Executive Summary", "Admin Panel", "Profile"])
else:
    page = st.sidebar.radio("Navigation", ["Dashboard", "Advanced Analytics", "Executive Summary", "Profile"])
# ===================================================
# TAB NAVIGATION DROPDOWNS FOR ALL PAGES (Auto-navigate)
# ===================================================
st.sidebar.markdown("---")
st.sidebar.markdown("### 🗂️ Quick Tab Navigation")

# Dashboard Tabs Dropdown
if page == "Dashboard":
    st.sidebar.markdown("#### 📊 Dashboard Tabs")
    dashboard_tabs = {
        "📋 Stock Status Table": 0,
        "📈 KPIs & Analytics": 1,
        "💡 Decision Briefs": 2,
        "📍 Hubs Distribution": 3,
        "📦 Supply Planning": 4,
        "📋 Purchase Order Status": 5,
        "🚚 New Deliveries": 6
    }

    selected_dashboard_tab = st.sidebar.selectbox(
        "Select tab to navigate:",
        list(dashboard_tabs.keys()),
        key="dashboard_tab_nav"
    )

    if selected_dashboard_tab:
        target_index = dashboard_tabs[selected_dashboard_tab]
        if st.session_state.get('last_dashboard_tab') != target_index:
            st.session_state.last_dashboard_tab = target_index
            st.session_state.go_to_dashboard_tab = target_index
            st.rerun()

# Advanced Analytics Tabs Dropdown
elif page == "Advanced Analytics":
    st.sidebar.markdown("#### 📊 Advanced Analytics Tabs")
    analytics_tabs = {
        "🏆 Branch Ranking": 0,
        "🔄 Redistribution": 1,
        "📧 Critical Alerts": 2,
        "⏰ Expiry Notifications": 3,
        "📊 Program Comparison": 4,
        "🗺️ Regional Map": 5,
        "👁️ Popular Materials": 6,
        "👥 User Analytics": 7,
        "📅 Report Scheduling": 8
    }

    selected_analytics_tab = st.sidebar.selectbox(
        "Select tab to navigate:",
        list(analytics_tabs.keys()),
        key="analytics_tab_nav"
    )

    # Auto-navigate when selection changes
    if selected_analytics_tab:
        target_index = analytics_tabs[selected_analytics_tab]
        if st.session_state.get('last_analytics_tab') != target_index:
            st.session_state.last_analytics_tab = target_index
            st.session_state.go_to_analytics_tab = target_index
            st.rerun()

# Executive Summary Sections Dropdown
elif page == "Executive Summary":
    st.sidebar.markdown("#### 📑 Executive Summary Sections")
    summary_sections = {
        "🎯 Performance Metrics": 0,
        "📊 Stock Status": 1,
        "⚠️ Risk Summary": 2,
        "📍 Hubs Distribution": 3,
        "🏆 Branch Ranking": 4,
        "📊 Program Performance": 5
    }

    selected_summary_section = st.sidebar.selectbox(
        "Select section to navigate:",
        list(summary_sections.keys()),
        key="summary_tab_nav"
    )

    # Auto-navigate when selection changes
    if selected_summary_section:
        target_index = summary_sections[selected_summary_section]
        if st.session_state.get('last_summary_section') != target_index:
            st.session_state.last_summary_section = target_index
            st.session_state.go_to_summary_section = target_index
            st.rerun()

st.sidebar.markdown("---")
# ---------------------------------------------------
# Data Refresh Controls
# ---------------------------------------------------
st.sidebar.divider()
st.sidebar.markdown("### 🔄 Data Updates")
st.sidebar.caption(f"📅 Data as of: {st.session_state.data_timestamp.strftime('%Y-%m-%d %H:%M:%S')}")

if st.session_state['user']['role'] == 'admin':
    st.sidebar.info("📊 Data Sources:\n- Supabase: Real-time\n- Google Sheets (AMC/Pipeline): Cached 5 min\n- Google Sheets (Branch AMC): Cached 5 min")

auto_refresh = st.sidebar.checkbox("Auto-refresh every 5 minutes", value=st.session_state.auto_refresh)
if auto_refresh != st.session_state.auto_refresh:
    st.session_state.auto_refresh = auto_refresh

if st.session_state.auto_refresh:
    time_since_refresh = (datetime.now() - st.session_state.data_timestamp).total_seconds()
    if time_since_refresh > 300:
        st.cache_data.clear()
        st.session_state.data_timestamp = datetime.now()
        st.rerun()

if st.sidebar.button("🔄 Refresh Now", use_container_width=True, type="primary"):
    st.cache_data.clear()
    st.session_state.data_timestamp = datetime.now()
    st.rerun()

if st.session_state['user']['role'] == 'admin':
    if st.sidebar.button("🗑️ Clear All Caches", use_container_width=True):
        st.cache_data.clear()
        st.cache_resource.clear()
        st.session_state.data_timestamp = datetime.now()
        st.success("All caches cleared! Refreshing...")
        time.sleep(1)
        st.rerun()

# ---------------------------------------------------
# Admin Panel - Supabase Management
# ---------------------------------------------------
if st.session_state['user']['role'] == 'admin':
    with st.sidebar.expander("📁 Supabase Management"):
        st.caption("Manage data in Supabase")

        table_info = get_table_info()
        if table_info:
            st.info(f"Current data: {table_info['rows']} rows, {table_info['columns']} columns")

        st.markdown("#### Upload New Data")
        st.caption("Upload Excel or CSV file with your data")

        uploaded_file = st.file_uploader("Choose Excel or CSV file", type=['xlsx', 'csv', 'xls'], key='data_upload')

        if uploaded_file:
            try:
                if uploaded_file.name.endswith('.csv'):
                    df_upload = pd.read_csv(uploaded_file)
                else:
                    df_upload = pd.read_excel(uploaded_file)

                st.write(f"Preview of data to upload ({len(df_upload)} rows):")
                st.dataframe(df_upload.head())

                col1, col2 = st.columns(2)
                with col1:
                    if st.button("📤 Upload to Supabase", use_container_width=True, type="primary"):
                        with st.spinner("Validating and uploading..."):
                            if validate_upload_data(df_upload):
                                if upload_to_supabase(df_upload):
                                    st.cache_data.clear()
                                    st.session_state.data_timestamp = datetime.now()
                                    st.success("Upload successful! Refreshing...")
                                    time.sleep(1)
                                    st.rerun()
                with col2:
                    if st.button("⚠️ Clear All Data", use_container_width=True):
                        with st.spinner("Clearing Supabase data..."):
                            if clear_supabase_table():
                                st.cache_data.clear()
                                st.session_state.data_timestamp = datetime.now()
                                st.success("Data cleared! Refreshing...")
                                time.sleep(1)
                                st.rerun()
            except Exception as e:
                st.error(f"Error reading file: {e}")

        st.markdown("---")
        st.caption("Data stored in Supabase table: `health_data`")

# ---------------------------------------------------
# Logout
# ---------------------------------------------------
st.sidebar.divider()
if st.sidebar.button("🚪 Logout", use_container_width=True):
    st.session_state['auth'] = False
    st.session_state['user'] = None
    st.rerun()

# ---------------------------------------------------
# Route to pages
# ---------------------------------------------------
if page == "Profile":
    show_profile_page()
    st.stop()
elif page == "Admin Panel" and st.session_state['user']['role'] == 'admin':
    show_admin_panel()
    st.stop()
elif page == "Advanced Analytics":
    # ===================================================
    # ADVANCED ANALYTICS PAGE (9 TABS - WITH HMOS CHANGES)
    # ===================================================
    st.markdown("<h1 style='font-size: 32px; font-weight: bold; font-family: Times New Roman;' class='gradient-text'>Advanced Analytics Dashboard</h1>", unsafe_allow_html=True)
if 'go_to_dashboard_tab' not in st.session_state:
    st.session_state.go_to_dashboard_tab = None
if 'go_to_analytics_tab' not in st.session_state:
    st.session_state.go_to_analytics_tab = None
if 'go_to_summary_section' not in st.session_state:
    st.session_state.go_to_summary_section = None
if 'last_dashboard_tab' not in st.session_state:
    st.session_state.last_dashboard_tab = None
if 'last_analytics_tab' not in st.session_state:
    st.session_state.last_analytics_tab = None
if 'last_summary_section' not in st.session_state:
    st.session_state.last_summary_section = None

# Auto-navigate to selected analytics tab
if st.session_state.go_to_analytics_tab is not None:
    tab_index = st.session_state.go_to_analytics_tab
    st.session_state.go_to_analytics_tab = None

    st.components.v1.html(f"""
    <script>
        setTimeout(function() {{
            var tabs = window.parent.document.querySelectorAll('button[data-baseweb="tab"]');
            if (tabs && tabs[{tab_index}]) {{
                tabs[{tab_index}].click();
            }}
        }}, 100);
    </script>
    """, height=0)
    # Create 9 sub-tabs within Advanced Analytics
    aa_tab1, aa_tab2, aa_tab3, aa_tab4, aa_tab5, aa_tab6, aa_tab7, aa_tab8, aa_tab9 = st.tabs([
        "🏆 Branch Ranking", "🔄 Redistribution", "📧 Critical Alerts", "⏰ Expiry Notifications",
        "📊 Program Comparison", "🗺️ Regional Map", "👁️ Popular Materials", "👥 User Analytics", 
        "📅 Report Scheduling"
    ])

    # ========== TAB 1: Branch Ranking (WITH HMOS) ==========
    with aa_tab1:
        st.markdown("<h3 style='font-size: 24px; font-weight: bold;'>Branch Ranking</h3>", unsafe_allow_html=True)
        st.caption("Availability = materials with HMOS ≥ 0.5 (at least 2 weeks of stock)")
        st.caption("Ranking based on: Availability (100%=1), Avg HMOS (2-4 months=1), Stock-out Materials (0=1)")

        if branch_amc_data is not None and not branch_amc_data.empty:
            branch_stock_cols = [col for col in df.columns if 'Branch' in col and col != 'Material Description']
            amc_branch_cols = [col for col in branch_amc_data.columns if col != 'Material Description']
            rankings = []

            for amc_branch in amc_branch_cols:
                if amc_branch == 'Material Description':
                    continue

                stock_col = None
                for bc in branch_stock_cols:
                    if amc_branch == bc:
                        stock_col = bc
                        break

                if stock_col:
                    try:
                        merged = pd.merge(
                            df[['Material Description', stock_col]], 
                            branch_amc_data[['Material Description', amc_branch]], 
                            on='Material Description', how='inner'
                        )

                        if not merged.empty:
                            stock_values = pd.to_numeric(merged[stock_col + '_x'], errors='coerce').fillna(0)
                            amc_values = pd.to_numeric(merged[amc_branch + '_y'], errors='coerce').fillna(1)
                            amc_values = amc_values.replace(0, 1)
                            branch_hmos = (stock_values / amc_values).values

                            availability_count = np.sum(branch_hmos >= 0.5)
                            total_materials = len(branch_hmos)
                            availability_score = (availability_count / total_materials * 100) if total_materials > 0 else 0
                            valid_hmos = branch_hmos[branch_hmos > 0]
                            avg_hmos = np.mean(valid_hmos) if len(valid_hmos) > 0 else 0
                            stock_out_materials = total_materials - availability_count

                            rankings.append({
                                'Branch': amc_branch,
                                'Availability Score (%)': round(availability_score, 1),
                                'Average HMOS': round(avg_hmos, 2),
                                'Stock-out Materials': stock_out_materials,
                                'Total Materials': total_materials
                            })
                    except Exception:
                        continue

            if rankings:
                for item in rankings:
                    composite_score = 0
                    if item['Availability Score (%)'] == 100:
                        composite_score += 1
                    if 2 <= item['Average HMOS'] <= 4:
                        composite_score += 1
                    if item['Stock-out Materials'] == 0:
                        composite_score += 1
                    item['Composite Score'] = composite_score

                rankings_df = pd.DataFrame(rankings).sort_values(['Composite Score', 'Availability Score (%)'], ascending=[False, False])
                rankings_df['Rank'] = range(1, len(rankings_df) + 1)
                rankings_df = rankings_df.drop(columns=['Composite Score'])

                col1, col2 = st.columns(2)
                with col1:
                    st.metric("🏆 Top Branch", rankings_df.iloc[0]['Branch'])
                    st.metric("📊 Best Availability Score", f"{rankings_df.iloc[0]['Availability Score (%)']}%")
                with col2:
                    st.metric("📉 Bottom Branch", rankings_df.iloc[-1]['Branch'])
                    st.metric("⚠️ Worst Availability Score", f"{rankings_df.iloc[-1]['Availability Score (%)']}%")

                st.dataframe(rankings_df, use_container_width=True, hide_index=True)

                fig = px.bar(rankings_df, x='Branch', y='Availability Score (%)', color='Availability Score (%)',
                            color_continuous_scale='RdYlGn', title='Branch Availability Scores (HMOS ≥ 0.5)')
                fig.update_layout(height=500, xaxis_tickangle=-45)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No branch data available for ranking")
        else:
            st.info("Branch AMC data not available")

    # ========== TAB 2: Redistribution Recommendations (WITH HMOS) ==========
    with aa_tab2:
        st.markdown("<h3 style='font-size: 24px; font-weight: bold;'>Stock Redistribution Recommendations</h3>", unsafe_allow_html=True)
        st.info("""
        **Redistribution Criteria:**
        - Trigger: Expiry Risk **OR** Understock (National NMOS < 6) **OR** Branch HMOS > 8
        - Source Branch: HMOS > 8
        - Target Branch: HMOS < 0.5
        - Target HMOS after transfer: 5 months
        """)

        if branch_amc_data is not None and not branch_amc_data.empty:
            branch_stock_cols = [col for col in df.columns if 'Branch' in col and col != 'Material Description']
            amc_branch_cols = [col for col in branch_amc_data.columns if col != 'Material Description']
            recommendations = []

            df_aligned = df.set_index('Material Description')
            amc_aligned = branch_amc_data.set_index('Material Description')
            common_materials = df_aligned.index.intersection(amc_aligned.index)

            for material in list(common_materials):
                material_row = df[df['Material Description'] == material]
                if material_row.empty:
                    continue

                has_expiry_risk = material_row.iloc[0].get('Has Expiry Risk', False)
                national_nmos = material_row.iloc[0].get('NMOS', 10)
                if pd.isna(national_nmos):
                    national_nmos = 10

                is_understock = national_nmos < 6
                trigger = has_expiry_risk or is_understock

                if not trigger:
                    continue

                branch_hmos = {}
                branch_stock = {}
                branch_amc_val = {}

                for amc_branch in amc_branch_cols:
                    stock_col = None
                    for bc in branch_stock_cols:
                        if amc_branch == bc:
                            stock_col = bc
                            break

                    if stock_col and stock_col in df_aligned.columns:
                        try:
                            stock = pd.to_numeric(df_aligned.loc[material, stock_col], errors='coerce')
                            amc = pd.to_numeric(amc_aligned.loc[material, amc_branch], errors='coerce')
                            if pd.isna(stock):
                                stock = 0
                            if pd.isna(amc) or amc <= 0:
                                amc = 1
                            hmos = stock / amc
                            branch_hmos[amc_branch] = hmos
                            branch_stock[amc_branch] = stock
                            branch_amc_val[amc_branch] = amc
                        except Exception:
                            continue

                overstocked = [b for b, hmos in branch_hmos.items() if hmos > 8]
                understocked = [b for b, hmos in branch_hmos.items() if 0 < hmos < 0.5]

                for source in overstocked:
                    for target in understocked:
                        if source != target:
                            excess = (branch_hmos[source] - 5) * branch_amc_val[source]
                            deficit = (5 - branch_hmos[target]) * branch_amc_val[target]
                            transfer_qty = min(max(0, excess), max(0, deficit))

                            if transfer_qty > 0 and transfer_qty >= branch_amc_val[target]:
                                if has_expiry_risk and is_understock:
                                    priority = "URGENT"
                                elif has_expiry_risk or is_understock:
                                    priority = "HIGH"
                                else:
                                    priority = "MEDIUM"

                                recommendations.append({
                                    'Material': material,
                                    'Trigger': ('Expiry Risk' if has_expiry_risk else '') + (' + Understock' if is_understock else ''),
                                    'Source Branch': source,
                                    'Target Branch': target,
                                    'Source HMOS': round(branch_hmos[source], 2),
                                    'Target HMOS': round(branch_hmos[target], 2),
                                    'Recommended Transfer Qty': int(transfer_qty),
                                    'Priority': priority
                                })

            if recommendations:
                st.dataframe(pd.DataFrame(recommendations), use_container_width=True, hide_index=True)
                total_qty = sum(r['Recommended Transfer Qty'] for r in recommendations)
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("🔄 Total Transfer Opportunities", len(recommendations))
                with col2:
                    st.metric("📦 Total Recommended Transfer", f"{int(total_qty):,} units")
                with col3:
                    urgent = len([r for r in recommendations if r['Priority'] == 'URGENT'])
                    st.metric("⚠️ Urgent Transfers", urgent)
                st.download_button("📥 Download Redistribution Plan", pd.DataFrame(recommendations).to_csv(index=False), "redistribution_plan.csv")
            else:
                st.success("✅ No redistribution opportunities identified. No materials meet the criteria.")
        else:
            st.info("Branch AMC data not available")

    # ========== TAB 3: Critical Alerts ==========
    with aa_tab3:
        st.markdown("<h3 style='font-size: 24px; font-weight: bold;'>Critical Stock-Out Alerts</h3>", unsafe_allow_html=True)

        critical_items = []
        for idx, row in df_filtered.iterrows():
            nmos = row.get('NMOS', 0)
            if pd.notna(nmos) and nmos < 0.5:
                rec = get_stock_out_recommendation(row)
                critical_items.append({
                    'Material': row['Material Description'],
                    'NMOS': round(nmos, 2),
                    'NSOH': row.get('NSOH', 0),
                    'AMC': row.get('AMC', 0),
                    'Stock Status': 'CRITICAL STOCK OUT',
                    'Recommendation': rec
                })
            elif pd.notna(nmos) and nmos < 1:
                rec = get_stock_out_recommendation(row)
                critical_items.append({
                    'Material': row['Material Description'],
                    'NMOS': round(nmos, 2),
                    'NSOH': row.get('NSOH', 0),
                    'AMC': row.get('AMC', 0),
                    'Stock Status': 'STOCK OUT',
                    'Recommendation': rec
                })

        if critical_items:
            critical_df = pd.DataFrame(critical_items)
            st.error(f"⚠️ {len(critical_items)} critical items requiring immediate attention!")
            col1, col2 = st.columns(2)
            with col1:
                critical_count = len([c for c in critical_items if c['Stock Status'] == 'CRITICAL STOCK OUT'])
                st.metric("🔴 Critical Stock-Outs", critical_count, delta="URGENT", delta_color="inverse")
            with col2:
                stockout_count = len([c for c in critical_items if c['Stock Status'] == 'STOCK OUT'])
                st.metric("🟡 Stock Out", stockout_count)
            st.dataframe(critical_df, use_container_width=True, hide_index=True)
            if st.button("📧 Send Email Alerts (Simulated)"):
                st.success(f"✅ Email alert would be sent to supply chain team with {len(critical_items)} critical items")
        else:
            st.success("✅ No critical stock-outs detected. All materials have adequate stock levels!")

    # ========== TAB 4: Expiry Notifications (CORRECTED) ==========
    with aa_tab4:
        st.markdown("<h3 style='font-size: 24px; font-weight: bold;'>Expiring Stock Notifications</h3>", unsafe_allow_html=True)

        notifications = []
        current_date = datetime.now()

        for idx, row in df_filtered.iterrows():
            expiry_str = row.get('Expiry', '')
            if pd.isna(expiry_str) or expiry_str == '':
                continue

            amc = row.get('AMC', 0)
            try:
                amc = float(amc) if pd.notna(amc) else 0
            except:
                amc = 0

            if amc <= 0:
                continue

            pattern = r'(\d[\d,]*)\s*\(([A-Za-z]+)-(\d{4})\)'
            matches = re.findall(pattern, str(expiry_str))
            month_map = {'Jan':1, 'Feb':2, 'Mar':3, 'Apr':4, 'May':5, 'Jun':6,
                        'Jul':7, 'Aug':8, 'Sep':9, 'Oct':10, 'Nov':11, 'Dec':12}

            for qty_str, month, year in matches:
                try:
                    qty = float(qty_str.replace(',', ''))
                    month_num = month_map.get(month[:3], 1)
                    expiry_date = datetime(int(year), month_num, 1)
                    months_to_expiry = (expiry_date.year - current_date.year) * 12 + (expiry_date.month - current_date.month)

                    months_of_stock_this_batch = qty / amc

                    if months_of_stock_this_batch <= months_to_expiry:
                        continue

                    if months_to_expiry <= 3:
                        priority = "🔴 CRITICAL"
                    elif months_to_expiry <= 6:
                        priority = "🟡 HIGH"
                    elif months_to_expiry <= 12:
                        priority = "🔵 MEDIUM"
                    else:
                        continue

                    rec = get_expiry_risk_recommendation(row)

                    notifications.append({
                        'Material': row['Material Description'],
                        'Priority': priority,
                        'Message': f"{int(qty):,} units expiring in {months_to_expiry} month(s)",
                        'Recommendation': rec,
                        'Quantity': int(qty),
                        'Expiry Date': expiry_date.strftime('%b-%Y'),
                        'Months Left': months_to_expiry,
                        'Months to Consume': round(months_of_stock_this_batch, 1)
                    })
                except:
                    continue

        if notifications:
            notif_df = pd.DataFrame(notifications).sort_values('Months Left')
            st.warning(f"⚠️ {len(notifications)} items with expiring stock that cannot be fully consumed before expiry")

            priority_counts = notif_df['Priority'].value_counts()
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("🔴 Critical (≤3 months)", priority_counts.get("🔴 CRITICAL", 0))
            with col2:
                st.metric("🟡 High (≤6 months)", priority_counts.get("🟡 HIGH", 0))
            with col3:
                st.metric("🔵 Medium (≤12 months)", priority_counts.get("🔵 MEDIUM", 0))

            st.dataframe(notif_df[['Material', 'Priority', 'Message', 'Recommendation', 'Quantity', 'Expiry Date', 'Months to Consume']], 
                        use_container_width=True, hide_index=True)

            if st.button("📢 Add to Dashboard Notifications"):
                for _, row in notif_df.head(5).iterrows():
                    st.session_state.notifications.append(f"{row['Priority']}: {row['Material']} - {row['Message']}")
                st.success(f"Added {min(5, len(notifications))} notifications to sidebar")
                st.rerun()
        else:
            st.success("✅ No expiring stock detected. All expiring batches can be consumed before their expiry date!")

                # ========== TAB 5: Program Comparison ==========
    with aa_tab5:
        st.markdown("<h3 style='font-size: 24px; font-weight: bold;'>Program Performance Comparison</h3>", unsafe_allow_html=True)
        st.caption("Ranking based on: Availability (100%=1), SAP (≥65%=1), Stock-out Rate (0%=1), Overstock Rate (0%=1), Avg NMOS (6-18 months=1)")

        if google_sheets and len(google_sheets) > 0:
            program_metrics = []

            for program_name, program_df in google_sheets.items():
                if program_df.empty:
                    continue

                if 'Material Description' in program_df.columns and 'Material Description' in df.columns:
                    merged = program_df[['Material Description']].merge(
                        df[['Material Description', 'NMOS']], on='Material Description', how='left'
                    )
                    nmos_values = pd.to_numeric(merged['NMOS'], errors='coerce').dropna()

                    if len(nmos_values) > 0:
                        availability = (nmos_values > 1).mean() * 100
                        sap = ((nmos_values >= 6) & (nmos_values <= 18)).mean() * 100
                        stock_out_rate = (nmos_values < 1).mean() * 100
                        overstock_rate = (nmos_values > 18).mean() * 100
                        avg_nmos = nmos_values.mean()
                    else:
                        availability = sap = stock_out_rate = overstock_rate = 0
                        avg_nmos = 0

                    program_metrics.append({
                        'Program': program_name,
                        'Availability (%)': round(availability, 1),
                        'SAP Achievement (%)': round(sap, 1),
                        'Stock-out Rate (%)': round(stock_out_rate, 1),
                        'Overstock Rate (%)': round(overstock_rate, 1),
                        'Avg NMOS': round(avg_nmos, 2),
                        'Total Materials': len(merged)
                    })

            if program_metrics:
                for item in program_metrics:
                    composite_score = 0
                    if item['Availability (%)'] == 100:
                        composite_score += 1
                    if item['SAP Achievement (%)'] >= 65:
                        composite_score += 1
                    if item['Stock-out Rate (%)'] == 0:
                        composite_score += 1
                    if item['Overstock Rate (%)'] == 0:
                        composite_score += 1
                    if 6 <= item['Avg NMOS'] <= 18:
                        composite_score += 1
                    item['Composite Score'] = composite_score

                comparison_df = pd.DataFrame(program_metrics).sort_values(['Composite Score', 'Availability (%)'], ascending=[False, False])
                comparison_df['Rank'] = range(1, len(comparison_df) + 1)
                comparison_df = comparison_df.drop(columns=['Composite Score'])

                top_program = comparison_df.iloc[0]['Program']
                top_availability = comparison_df.iloc[0]['Availability (%)']
                bottom_program = comparison_df.iloc[-1]['Program']
                bottom_availability = comparison_df.iloc[-1]['Availability (%)']

                # Side by side cards for TOP and BOTTOM programs
                col_left, col_right = st.columns(2)

                with col_left:
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); 
                                border-radius: 15px; 
                                padding: 15px; 
                                text-align: center;
                                box-shadow: 0 5px 15px rgba(0,0,0,0.1);
                                animation: pulse 2s infinite;'>
                        <div style='font-size: 32px; margin-bottom: 5px;'>🏆</div>
                        <div style='font-size: 11px; text-transform: uppercase; letter-spacing: 1px; color: rgba(255,255,255,0.8);'>Top Program</div>
                        <div style='font-size: 18px; font-weight: bold; color: white; margin: 5px 0;'>{top_program}</div>
                        <div style='display: inline-block; background: rgba(255,255,255,0.2); border-radius: 50px; padding: 4px 12px;'>
                            <span style='font-size: 18px; font-weight: bold; color: white;'>{top_availability}%</span>
                            <span style='color: rgba(255,255,255,0.8); font-size: 11px;'> Availability</span>
                        </div>
                    </div>
                    <style>
                    @keyframes pulse {{
                        0% {{ transform: scale(1); }}
                        50% {{ transform: scale(1.02); }}
                        100% {{ transform: scale(1); }}
                    }}
                    </style>
                    """, unsafe_allow_html=True)

                with col_right:
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); 
                                border-radius: 15px; 
                                padding: 15px; 
                                text-align: center;
                                box-shadow: 0 5px 15px rgba(0,0,0,0.1);'>
                        <div style='font-size: 32px; margin-bottom: 5px;'>📉</div>
                        <div style='font-size: 11px; text-transform: uppercase; letter-spacing: 1px; color: rgba(255,255,255,0.8);'>Bottom Program</div>
                        <div style='font-size: 18px; font-weight: bold; color: white; margin: 5px 0;'>{bottom_program}</div>
                        <div style='display: inline-block; background: rgba(255,255,255,0.2); border-radius: 50px; padding: 4px 12px;'>
                            <span style='font-size: 18px; font-weight: bold; color: white;'>{bottom_availability}%</span>
                            <span style='color: rgba(255,255,255,0.8); font-size: 11px;'> Availability</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                st.dataframe(comparison_df, use_container_width=True, hide_index=True)

                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=comparison_df['Program'], 
                    y=comparison_df['Availability (%)'],
                    name='Availability (%)',
                    text=comparison_df['Availability (%)'].apply(lambda x: f'{x}%'),
                    textposition='outside',
                    marker_color='skyblue'
                ))
                fig.add_trace(go.Bar(
                    x=comparison_df['Program'], 
                    y=comparison_df['SAP Achievement (%)'],
                    name='SAP Achievement (%)',
                    text=comparison_df['SAP Achievement (%)'].apply(lambda x: f'{x}%'),
                    textposition='outside',
                    marker_color='lightgreen'
                ))
                fig.update_layout(
                    title='Key Metrics Comparison',
                    xaxis_title='Program',
                    yaxis_title='Percentage (%)',
                    barmode='group',
                    height=500,
                    xaxis_tickangle=-45
                )
                st.plotly_chart(fig, use_container_width=True)

                if len(comparison_df) >= 2:
                    categories = ['Availability (%)', 'SAP Achievement (%)', 'Stock-out Rate (%)', 'Overstock Rate (%)']
                    prog1 = comparison_df.iloc[0]
                    prog2 = comparison_df.iloc[1]

                    fig_radar = go.Figure()
                    fig_radar.add_trace(go.Scatterpolar(
                        r=[prog1[c] for c in categories], theta=categories, fill='toself', name=prog1['Program']
                    ))
                    fig_radar.add_trace(go.Scatterpolar(
                        r=[prog2[c] for c in categories], theta=categories, fill='toself', name=prog2['Program']
                    ))
                    fig_radar.update_layout(polar=dict(radialaxis=dict(visible=True, range=[0, 100])), 
                                           showlegend=True, title="Top 2 Programs Comparison", height=500)
                    st.plotly_chart(fig_radar, use_container_width=True)
            else:
                st.info("No program comparison data available")
        else:
            st.info("Google Sheets data not available for program comparison")

    # ========== TAB 6: Regional Map (WITH HMOS) ==========
    with aa_tab6:
        st.markdown("<h3 style='font-size: 24px; font-weight: bold;'>Regional Stock Distribution Map</h3>", unsafe_allow_html=True)
        st.caption("🔴 Red = HMOS < 2 (Understock) | 🟢 Green = HMOS 2-4 (Normal) | 🔵 Skyblue = HMOS > 4 (Overstock)")

        branch_coords = {
            'Adama Branch': [8.5483, 39.2696],
            'Addis Ababa Branch 1': [9.0320, 38.7469],
            'Addis Ababa Branch 2': [9.0320, 38.7469],
            'Arba Minch Branch': [6.0333, 37.5500],
            'Assosa Branch': [10.0667, 34.5333],
            'Bahir Dar Branch': [11.5742, 37.3613],
            'Dessie Branch': [11.1333, 39.6333],
            'Dire Dawa Branch': [9.6000, 41.8500],
            'Gambela Branch': [8.2500, 34.5833],
            'Gondar Branch': [12.6000, 37.4667],
            'Hawassa Branch': [7.0500, 38.4667],
            'Jigjiga Branch': [9.3500, 42.8000],
            'Jimma Branch': [7.6667, 36.8333],
            'Mekele Branch': [13.4967, 39.4769],
            'Shire Branch': [14.1000, 38.2833]
        }

        if branch_amc_data is not None and not branch_amc_data.empty:
            branch_stock_cols = [col for col in df.columns if 'Branch' in col and col != 'Material Description']
            map_data = []

            for stock_col in branch_stock_cols:
                if stock_col in branch_amc_data.columns:
                    merged = pd.merge(
                        df[['Material Description', stock_col]], 
                        branch_amc_data[['Material Description', stock_col]], 
                        on='Material Description', 
                        how='inner'
                    )

                    if not merged.empty:
                        stock_values = pd.to_numeric(merged[stock_col + '_x'], errors='coerce').fillna(0)
                        amc_values = pd.to_numeric(merged[stock_col + '_y'], errors='coerce').fillna(1)
                        amc_values = amc_values.replace(0, 1)
                        hmos_values = (stock_values / amc_values).values

                        valid_hmos = hmos_values[hmos_values > 0]
                        avg_hmos = np.mean(valid_hmos) if len(valid_hmos) > 0 else 0

                        if avg_hmos < 2:
                            status = "Understock"
                        elif avg_hmos <= 4:
                            status = "Normal"
                        else:
                            status = "Overstock"

                        coords = branch_coords.get(stock_col, [9.0, 38.0])

                        map_data.append({
                            'Branch': stock_col,
                            'Latitude': coords[0],
                            'Longitude': coords[1],
                            'Average HMOS': round(avg_hmos, 2),
                            'Status': status
                        })

            if map_data:
                map_df = pd.DataFrame(map_data)

                fig = px.scatter_mapbox(map_df, lat='Latitude', lon='Longitude', 
                                       size='Average HMOS', size_max=30,
                                       color='Status', hover_name='Branch', 
                                       hover_data=['Average HMOS'],
                                       color_discrete_map={'Understock': 'red', 'Normal': 'green', 'Overstock': 'skyblue'},
                                       zoom=5, height=600, title='Branch Stock Distribution Map (Average HMOS)')
                fig.update_layout(mapbox_style='open-street-map')
                fig.update_layout(margin=dict(l=0, r=0, t=30, b=0))
                st.plotly_chart(fig, use_container_width=True)
                st.dataframe(map_df[['Branch', 'Average HMOS', 'Status']], use_container_width=True, hide_index=True)
            else:
                st.info("Map data not available. No HMOS values could be calculated for the branches.")
        else:
            st.info("Branch AMC data not available for map")

    # ========== TAB 7: Popular Materials ==========
    with aa_tab7:
        st.markdown("<h3 style='font-size: 24px; font-weight: bold;'>Most Viewed Materials</h3>", unsafe_allow_html=True)
        st.caption("Tracks materials that users search for and view most frequently")

        if st.session_state.material_views:
            popular_df = pd.DataFrame([
                {'Material': k, 'Views': v} for k, v in st.session_state.material_views.items()
            ]).sort_values('Views', ascending=False).head(10)

            if not popular_df.empty:
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("👁️ Total Material Views", popular_df['Views'].sum())
                with col2:
                    st.metric("⭐ Most Viewed", popular_df.iloc[0]['Material'])

                fig = px.bar(popular_df, x='Material', y='Views', color='Views', title='Top 10 Most Viewed Materials')
                fig.update_layout(height=500, xaxis_tickangle=-45)
                st.plotly_chart(fig, use_container_width=True)
                st.dataframe(popular_df, use_container_width=True, hide_index=True)

                if st.button("🗑️ Reset View Tracking"):
                    st.session_state.material_views = {}
                    st.rerun()
            else:
                st.info("No material view data yet")
        else:
            st.info("No material view data yet. Start searching and viewing materials to see popularity analytics.")

    # ========== TAB 8: User Analytics ==========
    with aa_tab8:
        st.markdown("<h3 style='font-size: 24px; font-weight: bold;'>User Role Analytics</h3>", unsafe_allow_html=True)

        if st.session_state['user']['role'] == 'admin':
            if st.session_state.user_activity:
                activity_df = pd.DataFrame(st.session_state.user_activity)
                if not activity_df.empty:
                    role_summary = activity_df.groupby('role').agg({
                        'user': 'nunique',
                        'action': 'count'
                    }).rename(columns={'user': 'Unique Users', 'action': 'Total Actions'})
                    st.dataframe(role_summary, use_container_width=True)

                    top_users = activity_df.groupby('user').size().sort_values(ascending=False).head(5).reset_index()
                    top_users.columns = ['User', 'Activity Count']
                    st.subheader("👥 Most Active Users")
                    st.dataframe(top_users, use_container_width=True, hide_index=True)

                    if st.button("🗑️ Clear Activity Log"):
                        st.session_state.user_activity = []
                        st.rerun()
                else:
                    st.info("No user activity data yet")
            else:
                st.info("No user activity data yet. User actions will be tracked as they use the dashboard.")
        else:
            st.info("User analytics are only available to administrators.")
            if st.session_state.user_activity:
                my_activity = [a for a in st.session_state.user_activity if a.get('user') == st.session_state['user']['email']]
                st.metric("Your Activity Count", len(my_activity))

    # ========== TAB 9: Report Scheduling ==========
    with aa_tab9:
        st.markdown("<h3 style='font-size: 24px; font-weight: bold;'>Automated Report Scheduling</h3>", unsafe_allow_html=True)
        st.info("Configure automated reports to be sent via email on a schedule")

        col1, col2 = st.columns(2)
        with col1:
            report_type = st.selectbox("Report Type", ["Weekly Summary", "Monthly Full Report", "Quarterly Trend Analysis"])
            schedule_day = st.selectbox("Schedule Day", ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"])
        with col2:
            recipient_email = st.text_input("Recipient Email", value=st.session_state['user']['email'])
            schedule_time = st.time_input("Schedule Time", value=datetime.now().replace(hour=8, minute=0))

        st.markdown("### Report Content Options")
        include_sections = st.multiselect("Include Sections", 
                                          ["Stock Status", "KPIs & Analytics", "Decision Briefs", 
                                           "Branch Rankings", "Redistribution Recommendations", "Expiry Alerts"],
                                          default=["Stock Status", "KPIs & Analytics"])

        if st.button("📅 Schedule Report", use_container_width=True, type="primary"):
            st.success(f"✅ Report scheduled! {report_type} will be sent to {recipient_email} every {schedule_day} at {schedule_time.strftime('%H:%M')}")
            st.info("📧 In production, this would connect to a cron job / scheduler to automatically generate and email reports")

            st.markdown("### 📄 Report Preview")
            preview_df = df_filtered.head(10)[['Material Description', 'NMOS', 'Stock Status', 'Risk Type']].copy()
            st.dataframe(preview_df, use_container_width=True)
            st.download_button("📥 Download Sample Report", preview_df.to_csv(index=False), f"{report_type.replace(' ', '_')}_sample.csv", "text/csv")

        st.markdown("---")
        st.markdown("### 📧 Email Configuration (Admin Only)")
        if st.session_state['user']['role'] == 'admin':
            smtp_server = st.text_input("SMTP Server", placeholder="smtp.gmail.com")
            smtp_port = st.number_input("SMTP Port", value=587)
            sender_email = st.text_input("Sender Email")
            sender_password = st.text_input("Sender Password", type="password")

            if st.button("💾 Save Email Settings"):
                st.success("Email settings saved (simulated)")
        else:
            st.info("Contact administrator to configure email settings")

    st.stop()

elif page == "Executive Summary":
    # ===================================================
    # EXECUTIVE SUMMARY PAGE (UPDATED)
    # ===================================================
    st.markdown("<h1 style='font-size: 36px; font-weight: bold; font-family: Times New Roman;' class='gradient-text'>📊 Executive Summary Dashboard</h1>", unsafe_allow_html=True)
    st.caption(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Program: {sheet_name if sheet_name != 'All' else 'All Programs'}")

    # SECTION 1: PERFORMANCE METRICS (Colorful cards)
    st.markdown("---")
    st.markdown("<h2 style='font-size: 28px; font-weight: bold;'>🎯 1. Performance Metrics</h2>", unsafe_allow_html=True)

    if not df_filtered.empty and 'NMOS' in df_filtered.columns:
        nmos_values = pd.to_numeric(df_filtered['NMOS'], errors='coerce').dropna()

        availability = (nmos_values > 1).mean() * 100 if len(nmos_values) > 0 else 0
        sap_achievement = ((nmos_values >= 6) & (nmos_values <= 18)).mean() * 100 if len(nmos_values) > 0 else 0

        if 'Avail Gap' in df_filtered.columns:
            avail_gap_values = pd.to_numeric(df_filtered['Avail Gap'], errors='coerce').dropna()
            avg_avail_gap = avail_gap_values.mean() if len(avail_gap_values) > 0 else 0
        else:
            avg_avail_gap = 0

        # Colorful metric cards
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 15px; padding: 20px; color: white;'>
                <h3 style='margin:0; font-size: 14px; opacity:0.9'>AVAILABILITY</h3>
                <p style='font-size: 36px; font-weight: bold; margin:5px 0'>{availability:.1f}%</p>
                <p style='margin:0; font-size: 12px; opacity:0.8'>Target: 100%</p>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); border-radius: 15px; padding: 20px; color: white;'>
                <h3 style='margin:0; font-size: 14px; opacity:0.9'>SAP ACHIEVEMENT</h3>
                <p style='font-size: 36px; font-weight: bold; margin:5px 0'>{sap_achievement:.1f}%</p>
                <p style='margin:0; font-size: 12px; opacity:0.8'>Target: 65%</p>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); border-radius: 15px; padding: 20px; color: white;'>
                <h3 style='margin:0; font-size: 14px; opacity:0.9'>AVAIL. GAP</h3>
                <p style='font-size: 36px; font-weight: bold; margin:5px 0'>{avg_avail_gap:.1f}%</p>
                <p style='margin:0; font-size: 12px; opacity:0.8'>Hubs% - Head Office%</p>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # SECTION: STOCK STATUS (New title)
        st.markdown("<h2 style='font-size: 28px; font-weight: bold;'>📊 Stock Status</h2>", unsafe_allow_html=True)

        total_materials = len(df_filtered)
        stock_out_count = len(df_filtered[df_filtered['Stock Status'] == 'Stock Out'])
        understock_count = len(df_filtered[df_filtered['Stock Status'] == 'Understock'])
        normal_count = len(df_filtered[df_filtered['Stock Status'] == 'Normal Stock'])
        overstock_count = len(df_filtered[df_filtered['Stock Status'] == 'Overstock'])

        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("📦 Total Materials", total_materials)
        with col2:
            st.metric("🔴 Stock Out", stock_out_count)
        with col3:
            st.metric("🟡 Understock", understock_count)
        with col4:
            st.metric("🟢 Normal Stock", normal_count)
        with col5:
            st.metric("🔵 Overstock", overstock_count)

        # SECTION 2: RISK SUMMARY (Colorful table)
        st.markdown("---")
        st.markdown("<h2 style='font-size: 28px; font-weight: bold;'>⚠️ 2. Risk Summary</h2>", unsafe_allow_html=True)

        critical_risk = len(df_filtered[df_filtered['Risk Type'] == 'Critical Risk'])
        risk_stock_out = len(df_filtered[df_filtered['Risk Type'] == 'Risk of Stock out'])
        expiry_risk = len(df_filtered[df_filtered['Risk Type'] == 'Expiry Risk'])

        st.markdown(f"""
        <table style='width: 100%; border-collapse: collapse; margin: 20px 0;'>
            <tr style='background-color: #ff4444; color: white; text-align: center;'>
                <th style='padding: 15px; font-size: 18px; border-radius: 10px 0 0 0;'>🔴 CRITICAL RISK</th>
                <th style='padding: 15px; font-size: 18px; background-color: #ffa500;'>🟡 RISK OF STOCK OUT</th>
                <th style='padding: 15px; font-size: 18px; background-color: #ff9800; border-radius: 0 10px 0 0;'>⚠️ EXPIRY RISK</th>
            </tr>
            <tr style='text-align: center; background-color: #f8f9fa;'>
                <td style='padding: 20px; font-size: 42px; font-weight: bold; color: #ff4444;'>{critical_risk}</td>
                <td style='padding: 20px; font-size: 42px; font-weight: bold; color: #ffa500;'>{risk_stock_out}</td>
                <td style='padding: 20px; font-size: 42px; font-weight: bold; color: #ff9800;'>{expiry_risk}</td>
            </tr>
            <tr style='text-align: center; background-color: #ffffff;'>
                <td style='padding: 10px; font-size: 14px;'>require URGENT attention</td>
                <td style='padding: 10px; font-size: 14px;'>need expediting</td>
                <td style='padding: 10px; font-size: 14px;'>approaching expiration</td>
            </tr>
        </table>
        """, unsafe_allow_html=True)

        # SECTION 3: STOCK DISTRIBUTION ACROSS HUBS (unchanged)
        st.markdown("---")
        st.markdown("<h2 style='font-size: 28px; font-weight: bold;'>📍 3. Stock Distribution Across Hubs by MOS</h2>", unsafe_allow_html=True)

        if 'CV Category' in df_filtered.columns:
            cv_counts = df_filtered['CV Category'].value_counts()
            total_cv_materials = len(df_filtered[df_filtered['CV Category'] != 'Unknown'])

            low_variation = cv_counts.get('Low variation', 0)
            moderate_variation = cv_counts.get('Moderate variation', 0)
            high_variation = cv_counts.get('High variation', 0)

            low_pct = (low_variation / total_cv_materials * 100) if total_cv_materials > 0 else 0
            mod_pct = (moderate_variation / total_cv_materials * 100) if total_cv_materials > 0 else 0
            high_pct = (high_variation / total_cv_materials * 100) if total_cv_materials > 0 else 0

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("📊 Total Materials", total_cv_materials)
            with col2:
                st.metric("🟢 Low Variation (<50%)", f"{low_variation} ({low_pct:.1f}%)")
            with col3:
                st.metric("🟡 Moderate Variation (50-100%)", f"{moderate_variation} ({mod_pct:.1f}%)")
            with col4:
                st.metric("🔴 High Variation (>100%)", f"{high_variation} ({high_pct:.1f}%)")
        else:
            st.info("CV Category data not available")

        # SECTION 4: BRANCH RANKING (Colorful gradient cards - name only)
        st.markdown("---")
        st.markdown("<h2 style='font-size: 28px; font-weight: bold;'>🏆 4. Top and Bottom Branch Ranking</h2>", unsafe_allow_html=True)
        st.caption("Availability = materials with NMOS ≥ 0.5 (at least 2 weeks of stock)")

        if branch_amc_data is not None and not branch_amc_data.empty:
            branch_stock_cols = [col for col in df.columns if 'Branch' in col and col != 'Material Description']
            amc_branch_cols = [col for col in branch_amc_data.columns if col != 'Material Description']
            rankings = []

            for amc_branch in amc_branch_cols:
                stock_col = None
                for bc in branch_stock_cols:
                    if amc_branch == bc:
                        stock_col = bc
                        break

                if stock_col:
                    try:
                        merged = pd.merge(
                            df[['Material Description', stock_col]], 
                            branch_amc_data[['Material Description', amc_branch]], 
                            on='Material Description', how='inner'
                        )
                        if not merged.empty:
                            stock_values = pd.to_numeric(merged[stock_col + '_x'], errors='coerce').fillna(0)
                            amc_values = pd.to_numeric(merged[amc_branch + '_y'], errors='coerce').fillna(1)
                            amc_values = amc_values.replace(0, 1)
                            branch_nmos = stock_values / amc_values
                            availability_count = np.sum(branch_nmos >= 0.5)
                            total_materials_branch = len(branch_nmos)
                            availability_score = (availability_count / total_materials_branch * 100) if total_materials_branch > 0 else 0
                            rankings.append({
                                'Branch': amc_branch,
                                'Score': availability_score
                            })
                    except Exception:
                        continue

            if rankings:
                rankings_df = pd.DataFrame(rankings).sort_values('Score', ascending=False)
                top_branch = rankings_df.iloc[0]['Branch']
                bottom_branch = rankings_df.iloc[-1]['Branch']

                # Colorful gradient cards - name only
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); border-radius: 15px; padding: 20px; color: white; text-align: center;'>
                        <h3 style='margin:0; font-size: 16px; opacity:0.9'>🏆 TOP BRANCH</h3>
                        <p style='font-size: 28px; font-weight: bold; margin: 10px 0;'>{top_branch}</p>
                    </div>
                    """, unsafe_allow_html=True)
                with col2:
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); border-radius: 15px; padding: 20px; color: white; text-align: center;'>
                        <h3 style='margin:0; font-size: 16px; opacity:0.9'>📉 BOTTOM BRANCH</h3>
                        <p style='font-size: 28px; font-weight: bold; margin: 10px 0;'>{bottom_branch}</p>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.info("No branch ranking data available")
        else:
            st.info("Branch AMC data not available")

        # SECTION 5: PROGRAM PERFORMANCE (Colorful gradient cards - name only)
        st.markdown("---")
        st.markdown("<h2 style='font-size: 28px; font-weight: bold;'>📊 5. Program Performance</h2>", unsafe_allow_html=True)

        if google_sheets and len(google_sheets) > 0:
            program_metrics = []
            for program_name, program_df in google_sheets.items():
                if program_df.empty or 'Material Description' not in program_df.columns:
                    continue

                merged = program_df[['Material Description']].merge(
                    df[['Material Description', 'NMOS']], on='Material Description', how='left'
                )
                nmos_values = pd.to_numeric(merged['NMOS'], errors='coerce').dropna()

                if len(nmos_values) > 0:
                    # Calculate composite score for ranking
                    availability_prog = (nmos_values > 1).mean() * 100
                    sap_prog = ((nmos_values >= 6) & (nmos_values <= 18)).mean() * 100
                    stock_out_rate = (nmos_values < 1).mean() * 100
                    overstock_rate = (nmos_values > 18).mean() * 100
                    avg_nmos_prog = nmos_values.mean()

                    composite_score = 0
                    if availability_prog == 100:
                        composite_score += 1
                    if sap_prog >= 65:
                        composite_score += 1
                    if stock_out_rate == 0:
                        composite_score += 1
                    if overstock_rate == 0:
                        composite_score += 1
                    if 6 <= avg_nmos_prog <= 18:
                        composite_score += 1

                    program_metrics.append({
                        'Program': program_name,
                        'Composite Score': composite_score
                    })

            if program_metrics:
                prog_df = pd.DataFrame(program_metrics).sort_values('Composite Score', ascending=False)
                top_program = prog_df.iloc[0]['Program']
                bottom_program = prog_df.iloc[-1]['Program']

                # Colorful gradient cards - name only
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); border-radius: 15px; padding: 20px; color: white; text-align: center;'>
                        <h3 style='margin:0; font-size: 16px; opacity:0.9'>🥇 TOP PROGRAM</h3>
                        <p style='font-size: 28px; font-weight: bold; margin: 10px 0;'>{top_program}</p>
                    </div>
                    """, unsafe_allow_html=True)
                with col2:
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); border-radius: 15px; padding: 20px; color: white; text-align: center;'>
                        <h3 style='margin:0; font-size: 16px; opacity:0.9'>📉 BOTTOM PROGRAM</h3>
                        <p style='font-size: 28px; font-weight: bold; margin: 10px 0;'>{bottom_program}</p>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.info("No program data available for comparison")
        else:
            st.info("Google Sheets data not available")

    # Download button
    st.markdown("---")

    summary_text = f"""
EXECUTIVE SUMMARY REPORT
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Program: {sheet_name if sheet_name != 'All' else 'All Programs'}

PERFORMANCE METRICS
- Availability: {availability:.1f}% (Target: 100%)
- SAP Achievement: {sap_achievement:.1f}% (Target: 65%)
- Avail. Gap: {avg_avail_gap:.1f}%

STOCK STATUS
- Total Materials: {total_materials}
- Stock Out: {stock_out_count}
- Understock: {understock_count}
- Normal Stock: {normal_count}
- Overstock: {overstock_count}

RISK SUMMARY
- Critical Risk: {critical_risk}
- Risk of Stock Out: {risk_stock_out}
- Expiry Risk: {expiry_risk}

STOCK DISTRIBUTION ACROSS HUBS
- Total Materials: {total_cv_materials}
- Low Variation: {low_variation} ({low_pct:.1f}%)
- Moderate Variation: {moderate_variation} ({mod_pct:.1f}%)
- High Variation: {high_variation} ({high_pct:.1f}%)

BRANCH RANKING
- Top Branch: {top_branch}
- Bottom Branch: {bottom_branch}

PROGRAM PERFORMANCE
- Top Program: {top_program}
- Bottom Program: {bottom_program}
"""

    st.download_button(
        label="📥 Download Executive Summary Report",
        data=summary_text,
        file_name=f"executive_summary_{datetime.now().strftime('%Y%m%d')}.txt",
        mime="text/plain",
        use_container_width=True
    )

    st.stop()

    # MAIN DASHBOARD (5 Tabs)
# ===================================================
else:
    # Header Row
    col_logo, col_title, col_settings = st.columns([1, 3, 1])

    with col_logo:
        # For local file - use st.image directly with file path
        try:
            st.image(r"C:\Users\BIYENSA.NEGERA\Desktop\Epss Logo .png", width=60)
        except:
            # Fallback if file not found
            st.markdown("<div style='background:#1a5276;border-radius:15px;padding:10px;text-align:center;width:60px'><p style='color:white;font-weight:bold;margin:0'>EPSS</p></div>", unsafe_allow_html=True)

    # Column 2: Title
    with col_title:
        st.markdown("""
        <div>
            <p style='font-size: 18px; color: #1a5276; margin: 0; font-weight: bold;'>ETHIOPIAN PHARMACEUTICAL SUPPLY SERVICE</p>
            <h1 style='font-size: 32px; font-weight: bold; font-family: Times New Roman; margin: 0;' class='gradient-text'>Health Program Medicines Dashboard</h1>
        </div>
        """, unsafe_allow_html=True)

    # Column 3: Display Settings
    with col_settings:
        st.markdown("""
        <div style='background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); border-radius: 15px; padding: 10px; border: 1px solid #dee2e6;'>
            <p style='margin: 0 0 5px 0; font-size: 12px; font-weight: bold; color: #1a5276; text-align: center;'>🎨 DISPLAY SETTINGS</p>
        </div>
        """, unsafe_allow_html=True)
        view_mode = st.selectbox("View Mode", ["Table View", "Card View"], index=0 if st.session_state.view_mode == "table" else 1, label_visibility="collapsed")
        st.session_state.view_mode = "table" if view_mode == "Table View" else "card"

    st.markdown("---")

    # Quick Summary Section with Program Name
    if not df_filtered.empty and 'NMOS' in df_filtered.columns:
        nmos_values = pd.to_numeric(df_filtered['NMOS'], errors='coerce').dropna()

        availability = (nmos_values > 1).mean() * 100 if len(nmos_values) > 0 else 0
        sap = ((nmos_values >= 6) & (nmos_values <= 18)).mean() * 100 if len(nmos_values) > 0 else 0

        if 'Avail Gap' in df_filtered.columns:
            avail_gap_values = pd.to_numeric(df_filtered['Avail Gap'], errors='coerce').dropna()
            avg_avail_gap = avail_gap_values.mean() if len(avail_gap_values) > 0 else 0
        else:
            avg_avail_gap = 0

        # Determine program name for Quick Summary title
        if sheet_name == "All":
            program_name = "All Programs"
        elif subcategory_filter != "All":
            program_name = f"{sheet_name} - {subcategory_filter}"
        else:
            program_name = sheet_name

        # Quick Summary Box with Program Name
        st.markdown(f"""
        <div style='background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); border-radius: 15px; padding: 15px; border: 1px solid #dee2e6; margin-bottom: 15px;'>
            <h4 style='margin: 0 0 10px 0; font-size: 16px; font-weight: bold; color: #1a5276;'>📊 {program_name} Quick Summary</h4>
        </div>
        """, unsafe_allow_html=True)

        # Row: Availability, SAP, Avail.Gap
        col_a, col_s, col_g = st.columns(3)
        with col_a:
            st.markdown(f"""
            <div style='text-align: center; padding: 10px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 10px; margin: 5px;'>
                <p style='margin: 0; font-size: 16px; color: white; opacity: 0.9'>AVAILABILITY</p>
                <p style='margin: 0; font-size: 24px; font-weight: bold; color: white;'>{availability:.1f}%</p>
            </div>
            """, unsafe_allow_html=True)
        with col_s:
            st.markdown(f"""
            <div style='text-align: center; padding: 10px; background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); border-radius: 10px; margin: 5px;'>
                <p style='margin: 0; font-size: 16px; color: white; opacity: 0.9'>SAP</p>
                <p style='margin: 0; font-size: 24px; font-weight: bold; color: white;'>{sap:.1f}%</p>
            </div>
            """, unsafe_allow_html=True)
        with col_g:
            st.markdown(f"""
            <div style='text-align: center; padding: 10px; background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); border-radius: 10px; margin: 5px;'>
                <p style='margin: 0; font-size: 16px; color: white; opacity: 0.9'>AVAIL. GAP</p>
                <p style='margin: 0; font-size: 24px; font-weight: bold; color: white;'>{avg_avail_gap:.1f}%</p>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("---")

    # Original 5-COLUMN QUICK STATS (keep as is)
    if not df_filtered.empty:
        col1, col2, col3, col4, col5 = st.columns(5)

        total_items = len(df_filtered)
        stock_out = len(df_filtered[df_filtered['Stock Status'] == 'Stock Out']) if 'Stock Status' in df_filtered.columns else 0
        understock = len(df_filtered[df_filtered['Stock Status'] == 'Understock']) if 'Stock Status' in df_filtered.columns else 0
        normal = len(df_filtered[df_filtered['Stock Status'] == 'Normal Stock']) if 'Stock Status' in df_filtered.columns else 0
        overstock = len(df_filtered[df_filtered['Stock Status'] == 'Overstock']) if 'Stock Status' in df_filtered.columns else 0

        if sheet_name == "All":
            program_display = "All Programs"
        elif subcategory_filter != "All":
            program_display = f"{sheet_name} - {subcategory_filter}"
        else:
            program_display = sheet_name

        with col1:
            st.metric(f"📊 {program_display} Total Items", total_items)
        with col2:
            st.metric("🔴 Stock Out", stock_out, delta=f"-{stock_out}" if stock_out > 0 else "0", delta_color="inverse")
        with col3:
            st.metric("🟡 Understock", understock, delta=f"-{understock}" if understock > 0 else "0", delta_color="inverse")
        with col4:
            st.metric("🟢 Normal Stock", normal, delta=f"+{normal}" if normal > 0 else "0", delta_color="normal")
        with col5:
            st.metric("🔵 Overstock", overstock, delta=f"-{overstock}" if overstock > 0 else "0", delta_color="inverse")

    # ---------------------------------------------------
        # Auto-navigate to selected dashboard tab
    if st.session_state.get('go_to_dashboard_tab') is not None:
        tab_index = st.session_state.go_to_dashboard_tab
        st.session_state.go_to_dashboard_tab = None

        st.components.v1.html(f"""
        <script>
            setTimeout(function() {{
                var tabs = window.parent.document.querySelectorAll('button[data-baseweb="tab"]');
                if (tabs && tabs[{tab_index}]) {{
                    tabs[{tab_index}].click();
                }}
            }}, 200);
        </script>
        """, height=0)
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "📋 Stock Status Table", 
    "📈 KPIs & Analytics", 
    "💡 Decision Briefs", 
    "📍 Hubs Distribution",
    "📦 Supply Planning",
    "📋 Purchase Order Status",
    "🚚 New Deliveries"
])

    # ---------------------------------------------------
# TAB 1 - Stock Status Table
# ---------------------------------------------------
with tab1:
    st.markdown("<h3 style='font-size: 28px; font-weight: bold; font-family: Times New Roman;'>Complete Stock Status Table</h3>", unsafe_allow_html=True)

    if not display_df_filtered.empty and 'Material Description' in display_df_filtered.columns:
        search_query = st.text_input(
            "Search by Material Description or any column value:",
            value=st.session_state.search_query,
            placeholder="Type to search... (e.g., 'artesunate', 'stock out', 'shipped')",
            key="search_input"
        )

        if search_query != st.session_state.search_query:
            st.session_state.search_query = search_query
            st.rerun()

        search_df = display_df_filtered.copy()

        if st.session_state.search_query:
            search_term = st.session_state.search_query.lower()
            string_df = search_df.astype(str)
            mask = string_df.apply(
                lambda col: col.str.lower().str.contains(search_term, na=False, regex=False)
            ).any(axis=1)
            search_df = search_df[mask]
            st.info(f"🔍 Found {len(search_df)} matching records for '{st.session_state.search_query}'")
            if st.button("🗑️ Clear Search"):
                st.session_state.search_query = ""
                st.rerun()
        else:
            st.info(f"📊 Showing all {len(search_df)} records")

        # Column ordering - keep original positions, just add DOS after NMOS
        cols = list(search_df.columns)
        if 'Material Description' in cols:
            cols.remove('Material Description')
            cols.insert(0, 'Material Description')
        if 'NMOS' in cols and 'AMC' in cols:
            cols.remove('NMOS')
            amc_index = cols.index('AMC') if 'AMC' in cols else 0
            cols.insert(amc_index + 1, 'NMOS')
        # Add DOS right after NMOS (keeping NMOS in original position after AMC)
        if 'DOS' in cols:
            cols.remove('DOS')
            nmos_index = cols.index('NMOS') if 'NMOS' in cols else 0
            cols.insert(nmos_index + 1, 'DOS')
        if 'Risk Type' in cols and 'Stock Status' in cols:
            cols.remove('Risk Type')
            status_index = cols.index('Stock Status') if 'Stock Status' in cols else 0
            cols.insert(status_index + 1, 'Risk Type')

        cols = [c for c in cols if c in search_df.columns]
        search_df = search_df[cols]

        if st.session_state.view_mode == "card":
            st.markdown("### 📇 Card View")
            cols_per_row = 3
            for i in range(0, len(search_df), cols_per_row):
                cols = st.columns(cols_per_row)
                for j, col in enumerate(cols):
                    idx = i + j
                    if idx < len(search_df):
                        row = search_df.iloc[idx]
                        status = row.get('Stock Status', '')
                        color_map = {
                            "Stock Out": "#ff4444",
                            "Understock": "#ffa500",
                            "Normal Stock": "#4CAF50",
                            "Overstock": "#2196F3"
                        }
                        border_color = color_map.get(status, "#ddd")
                        risk_type_value = row.get('Risk Type', 'N/A')
                        risk_color = {
                            "Critical Risk": "#9b59b6",
                            "Risk of Stock out": "#ffa500",
                            "Expiry Risk": "#ffa500"
                        }.get(risk_type_value, border_color)

                        material_desc = row.get('Material Description', 'N/A')
                        if pd.isna(material_desc) or material_desc is None:
                            material_desc = 'N/A'
                        else:
                            material_desc = str(material_desc)[:60]

                        dos_value = row.get('DOS', 0)
                        if pd.isna(dos_value):
                            dos_value = 0

                        with col:
                            st.markdown(f"""
                            <div class="stock-card" style='border-left: 4px solid {border_color};'>
                                <h4 style='color: {border_color}; margin-bottom: 10px;'>{material_desc}</h4>
                                <p><strong>📦 NSOH:</strong> {row.get('NSOH', 'N/A')}</p>
                                <p><strong>📈 AMC:</strong> {row.get('AMC', 'N/A')}</p>
                                <p><strong>⏰ NMOS:</strong> <span style='color: {border_color}; font-weight: bold;'>{row.get('NMOS', 'N/A')}</span></p>
                                <p><strong>📅 DOS:</strong> <span style='color: {"#ff4444" if dos_value > 0 else "#4CAF50"}; font-weight: bold;'>{dos_value} days</span></p>
                                <p><strong>📊 Status:</strong> {row.get('Stock Status', 'N/A')}</p>
                                <p><strong>🔄 TMOS:</strong> {row.get('TMOS', 'N/A')}</p>
                                <p><strong>⚠️ Risk Type:</strong> <span style='color: {risk_color}; font-weight: bold;'>{risk_type_value}</span></p>
                            </div>
                            """, unsafe_allow_html=True)
        else:
            def color_row(row):
                colors = {
                    "Stock Out": "background-color:red;color:white",
                    "Understock": "background-color:yellow",
                    "Normal Stock": "background-color:green;color:white",
                    "Overstock": "background-color:skyblue"
                }
                styles = [''] * len(row)
                for i, col in enumerate(row.index):
                    if col == 'Material Description':
                        styles[i] = colors.get(row['Stock Status'], '')
                        break
                return styles

            styled = search_df.style.apply(color_row, axis=1)

            def color_risk_type(val):
                if val == "Critical Risk":
                    return 'background-color: #9b59b6; color: white'
                elif val == "Risk of Stock out":
                    return 'background-color: #ffa500; color: white'
                elif val == "Expiry Risk":
                    return 'background-color: #ffa500; color: white'
                return ''

            styled = styled.map(color_risk_type, subset=['Risk Type'])

            def color_dos(val):
                if pd.notna(val) and val > 0:
                    return 'background-color: #ff4444; color: white; font-weight: bold'
                return ''

            if 'DOS' in search_df.columns:
                styled = styled.map(color_dos, subset=['DOS'])

            column_config = {
                "Material Description": st.column_config.TextColumn("Material Description", width=300, pinned=True),
                "Risk Type": st.column_config.TextColumn("Risk Type", width=150),
                "DOS": st.column_config.NumberColumn("DOS (Days)", width=100)
            }

            st.dataframe(styled, column_config=column_config, use_container_width=True, hide_index=True, height=min(800, (len(search_df) + 1) * 35))

            st.markdown("""
            <div class="legend-box">
                <h5 class="legend-title">📊 Color Legend - Stock Status:</h5>
                <div style="display: flex; gap: 20px; flex-wrap: wrap;">
                    <div class="legend-item"><div class="legend-color" style="background-color: red;"></div><span><strong>Red</strong> = Stock Out (NMOS &lt; 1)</span></div>
                    <div class="legend-item"><div class="legend-color" style="background-color: yellow;"></div><span><strong>Yellow</strong> = Understock (1 ≤ NMOS &lt; 6)</span></div>
                    <div class="legend-item"><div class="legend-color" style="background-color: green;"></div><span><strong>Green</strong> = Normal Stock (6 ≤ NMOS ≤ 18)</span></div>
                    <div class="legend-item"><div class="legend-color" style="background-color: skyblue;"></div><span><strong>Blue</strong> = Overstock (NMOS > 18)</span></div>
                    <div class="legend-item"><div class="legend-color" style="background-color: #9b59b6;"></div><span><strong>Purple</strong> = Critical Risk</span></div>
                    <div class="legend-item"><div class="legend-color" style="background-color: #ffa500;"></div><span><strong>Orange</strong> = Risk of Stock out or Expiry Risk</span></div>
                    <div class="legend-item"><div class="legend-color" style="background-color: #ff4444;"></div><span><strong>Red DOS</strong> = Days Out of Stock (>0 days)</span></div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        # DOWNLOAD REPORT
        st.markdown("---")
        st.markdown("<h4 style='font-size: 20px; font-weight: bold; font-family: Times New Roman;'>📥 Download Report</h4>", unsafe_allow_html=True)

        report_columns = []

        if 'Material Description' in df_filtered.columns:
            report_columns.append('Material Description')

        all_columns = list(df_filtered.columns)

        if 'Material Description' in all_columns and 'TMOS' in all_columns:
            mat_index = all_columns.index('Material Description')
            tmos_index = all_columns.index('TMOS')
            all_cols_between = all_columns[mat_index:tmos_index + 1]

            if 'AMC' in all_cols_between and 'NMOS' in all_cols_between:
                for col in all_cols_between:
                    if col == 'NMOS':
                        continue
                    report_columns.append(col)
                if 'AMC' in report_columns:
                    amc_pos = report_columns.index('AMC')
                    report_columns.insert(amc_pos + 1, 'NMOS')
                if 'DOS' in all_cols_between and 'DOS' not in report_columns:
                    nmos_pos = report_columns.index('NMOS') if 'NMOS' in report_columns else 0
                    report_columns.insert(nmos_pos + 1, 'DOS')
            else:
                report_columns = all_cols_between
        else:
            for col in all_columns:
                if col not in ['Stock Status', 'Risk of Stock', 'Hubs%', 'Head Office%', 'Avail Gap', 'CV (%)', 'CV Category', 'Has Expiry Risk', 'Expiry Risk Details', 'Assigned Subcategory', 'Risk Type']:
                    report_columns.append(col)

        if not report_columns:
            report_columns = ['Material Description']

        report_columns = list(dict.fromkeys(report_columns))

        report_df = df_filtered[report_columns].copy()

        for col in report_df.columns:
            if col in ['NMOS', 'GIT_MOS', 'LC_MOS', 'WB_MOS', 'TMD_MOS', 'TMOS']:
                report_df[col] = report_df[col].apply(lambda x: round(x, 2) if pd.notna(x) else "")
            elif col in ['NSOH', 'AMC', 'Hubs', 'Head Office']:
                report_df[col] = report_df[col].apply(lambda x: f"{int(x):,}" if pd.notna(x) and x != "" else "" if x == "" else x)
            elif col == 'DOS':
                report_df[col] = report_df[col].apply(lambda x: f"{int(x)} days" if pd.notna(x) and x > 0 else "0 days" if pd.notna(x) else "")

        output = BytesIO()

        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                if sheet_name == "All":
                    report_title = "All Programs Medicines Monthly National and Pipeline Report"
                else:
                    report_title = f"{sheet_name} Medicines Monthly National and Pipeline Report"

                report_df.to_excel(writer, sheet_name=report_title[:31], index=False)
                worksheet = writer.sheets[report_title[:31]]

                for column in report_df.columns:
                    try:
                        str_values = report_df[column].astype(str).replace('nan', '').replace('None', '')
                        if len(str_values) > 0:
                            column_length = max(str_values.map(len).max(), len(column))
                        else:
                            column_length = len(column)
                        column_length = min(column_length, 50)
                        col_idx = report_df.columns.get_loc(column)
                        col_letter = get_column_letter(col_idx + 1)
                        worksheet.column_dimensions[col_letter].width = column_length + 2
                    except Exception:
                        col_idx = report_df.columns.get_loc(column)
                        col_letter = get_column_letter(col_idx + 1)
                        worksheet.column_dimensions[col_letter].width = 15

            output.seek(0)
            st.download_button(
                label=f"📊 Download {sheet_name} Monthly National and Pipeline Report (Excel)",
                data=output,
                file_name=f"{sheet_name}_Monthly_National_Pipeline_Report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            st.caption(f"Report includes columns from Material Description to TMOS ({len(report_columns)} columns)")

        except Exception as excel_error:
            st.error(f"Error creating Excel file: {excel_error}")
            csv_data = report_df.to_csv(index=False)
            st.download_button(
                label=f"📊 Download {sheet_name} Monthly National and Pipeline Report (CSV)",
                data=csv_data,
                file_name=f"{sheet_name}_Monthly_National_Pipeline_Report_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )

                        # STOCK CHANGES TABLE
        st.markdown("---")
        st.markdown("<h4 style='font-size: 20px; font-weight: bold; font-family: Times New Roman;'>📈 Stock Changes Since Last Update</h4>", unsafe_allow_html=True)

        if st.session_state.nsoh_changes is not None and not st.session_state.nsoh_changes.empty:

            # Show summary metrics
            added = st.session_state.nsoh_changes[st.session_state.nsoh_changes['Change Type'] == '📦 STOCK ADDED']
            consumed = st.session_state.nsoh_changes[st.session_state.nsoh_changes['Change Type'] == '📉 STOCK CONSUMED']
            no_change = st.session_state.nsoh_changes[st.session_state.nsoh_changes['Change Type'] == '➖ NO CHANGE']

            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.metric("📋 Total Materials", len(st.session_state.nsoh_changes))
            with col2:
                st.metric("➕ Stock Added", len(added))
            with col3:
                st.metric("📉 Stock Consumed", len(consumed))
            with col4:
                st.metric("➖ No Change", len(no_change))
            with col5:
                # Calculate net change
                total_added = 0
                for val in added['NSOH_Change']:
                    try:
                        total_added += int(val.replace(',', '').replace('+', ''))
                    except:
                        pass
                total_consumed = 0
                for val in consumed['NSOH_Change']:
                    try:
                        total_consumed += abs(int(val.replace(',', '')))
                    except:
                        pass
                net_change = total_added - total_consumed
                st.metric("📊 Net Change", f"{net_change:+,}")

            st.markdown("---")

            # Add filter for change type
            filter_type = st.radio(
                "Filter by Change Type:",
                ["All", "📦 STOCK ADDED", "📉 STOCK CONSUMED", "➖ NO CHANGE"],
                horizontal=True
            )

            # Apply filter
            if filter_type != "All":
                filtered_changes = st.session_state.nsoh_changes[st.session_state.nsoh_changes['Change Type'] == filter_type]
            else:
                filtered_changes = st.session_state.nsoh_changes

            st.markdown(f"Showing **{len(filtered_changes)}** materials")

            # Single table showing all changes
            st.dataframe(
                filtered_changes,
                column_config={
                    "Material Description": st.column_config.TextColumn("Material", width=280),
                    "NSOH_previous": st.column_config.TextColumn("Previous NSOH", width=110),
                    "NSOH_now": st.column_config.TextColumn("Current NSOH", width=110),
                    "NSOH_Change": st.column_config.TextColumn("NSOH Change", width=110),
                    "NSOH_Change_Pct": st.column_config.TextColumn("Change %", width=90),
                    "NMOS_previous": st.column_config.TextColumn("Previous NMOS", width=100),
                    "NMOS_now": st.column_config.TextColumn("Current NMOS", width=100),
                    "NMOS_Change": st.column_config.TextColumn("NMOS Change", width=100),
                    "Change Type": st.column_config.TextColumn("Type", width=130),
                },
                use_container_width=True,
                hide_index=True
            )

            # Download button
            changes_csv = st.session_state.nsoh_changes.to_csv(index=False)
            st.download_button(
                label="📥 Download Stock Changes (CSV)",
                data=changes_csv,
                file_name=f"stock_changes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv",
                use_container_width=True
            )

        elif st.session_state.raw_previous_data is not None:
            st.info("ℹ️ No previous data available for comparison. Stock change tracking will appear after the next update.")
        else:
            st.info("ℹ️ Stock change tracking will appear after the next data update. (Need at least 2 data points to compare)")

    # ---------------------------------------------------
    # TAB 2 - KPIs & Analytics
    # ---------------------------------------------------
    with tab2:
        if sheet_name == "All":
            program_display = "All Programs"
        elif subcategory_filter != "All":
            program_display = f"{sheet_name} - {subcategory_filter}"
        else:
            program_display = sheet_name

        st.markdown(f"<h3 style='font-size: 28px; font-weight: bold; font-family: Times New Roman;'>{program_display} Medicines Performance Metrics</h3>", unsafe_allow_html=True)

        if not df_filtered.empty and 'NMOS' in df_filtered.columns:
            nmos_values = pd.to_numeric(df_filtered['NMOS'], errors='coerce').dropna()
            availability = (nmos_values > 1).mean() * 100 if len(nmos_values) > 0 else 0
            sap = ((nmos_values >= 6) & (nmos_values <= 18)).mean() * 100 if len(nmos_values) > 0 else 0

            if 'Avail Gap' in df_filtered.columns:
                avail_gap_values = pd.to_numeric(df_filtered['Avail Gap'], errors='coerce').dropna()
                avg_avail_gap = avail_gap_values.mean() if len(avail_gap_values) > 0 else 0
            else:
                avg_avail_gap = 0

            avail_delta = availability - 100

            if -5 <= avg_avail_gap <= 5:
                avail_gap_status = "✅ Within Target (-5% to 5%)"
                avail_gap_delta = None
            elif avg_avail_gap > 5:
                avail_gap_status = f"⚠️ Above Target (+{avg_avail_gap - 5:.1f}% over)"
                avail_gap_delta = f"+{avg_avail_gap - 5:.1f}%"
            else:
                avail_gap_status = f"⚠️ Below Target ({avg_avail_gap + 5:.1f}% under)"
                avail_gap_delta = f"{avg_avail_gap + 5:.1f}%"

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("📈 Availability", f"{availability:.1f}%", delta=f"{avail_delta:+.1f}%", delta_color="inverse")
                st.caption("Target: 100%")
            with col2:
                st.metric("🎯 SAP Achievement", f"{sap:.1f}%", delta=f"{sap - 65:.1f}%")
                st.caption("Target: 65%")
            with col3:
                st.metric("📊 Avail. Gap (Hubs% - Head Office%)", f"{avg_avail_gap:.1f}%", delta=avail_gap_delta, delta_color="inverse" if avg_avail_gap > 5 or avg_avail_gap < -5 else "off")
                st.caption(avail_gap_status)

            st.markdown("---")

            def create_kpi_fig(value, target, title):
                color = 'red' if value < target else 'black'
                return go.Figure(go.Indicator(mode="gauge+number", value=value, number={'suffix': '%', 'font': {'size': 36, 'color': color}}, title={'text': f"<b>{title}</b>", 'font': {'size': 24}}, gauge={'axis': {'range': [0, 100]}, 'bar': {'color': 'skyblue'}}))

            col1, col2 = st.columns(2)
            with col1:
                st.plotly_chart(create_kpi_fig(availability, 100, "Availability"), use_container_width=True)
            with col2:
                st.plotly_chart(create_kpi_fig(sap, 65, "SAP"), use_container_width=True)

            try:
                if 'Stock Status' in df_filtered.columns:
                    status_counts = df_filtered['Stock Status'].replace("", np.nan).dropna().value_counts()
                    if not status_counts.empty:
                        fig = px.pie(values=status_counts.values, names=status_counts.index, hole=0.5, color=status_counts.index, color_discrete_map={"Stock Out": "red", "Understock": "yellow", "Normal Stock": "green", "Overstock": "skyblue"})
                        fig.update_traces(textposition='inside', textinfo='percent+value', textfont_size=16)
                        total_count = len(df_filtered[df_filtered['Stock Status'] != ""])
                        fig.update_layout(title={'text': f"{program_display} Stock Status (Total: {total_count} items)", 'x': 0, 'xanchor': 'left', 'font': {'size': 20, 'weight': 'bold'}}, annotations=[dict(text=f"Total<br>{total_count}", x=0.5, y=0.5, font_size=20, showarrow=False)])
                        st.plotly_chart(fig, use_container_width=True)
            except Exception:
                pass

            st.markdown("### 🚚 Pipeline Status Analysis")
            pipeline_cols = [col for col in ['GIT_MOS', 'LC_MOS', 'WB_MOS', 'TMD_MOS'] if col in df_filtered.columns]
            if pipeline_cols:
                pipeline_summary = df_filtered[pipeline_cols].sum()
                if pipeline_summary.sum() > 0:
                    fig_pipeline = px.pie(values=pipeline_summary.values, names=pipeline_summary.index, title="Pipeline MOS Distribution", color_discrete_sequence=px.colors.sequential.Blues_r)
                    fig_pipeline.update_traces(textposition='inside', textinfo='percent+label')
                    st.plotly_chart(fig_pipeline, use_container_width=True)
                else:
                    st.info("No pipeline data available")
            else:
                st.info("Pipeline MOS columns not available")

            st.markdown("### ⚠️ Critical Items - Stock Out")
            stock_out_items = df_filtered[df_filtered['Stock Status'] == 'Stock Out'][['Material Description', 'NSOH', 'AMC', 'NMOS']].copy()
            if not stock_out_items.empty:
                stock_out_items = stock_out_items.sort_values('NMOS').head(10)
                fig_stock_out = go.Figure(data=[go.Bar(x=stock_out_items['Material Description'], y=stock_out_items['NMOS'], marker_color='red', text=stock_out_items['NMOS'].round(2), textposition='outside', name='NMOS')])
                fig_stock_out.update_layout(title="Top 10 Stock Out Items (Lowest NMOS)", xaxis_title="Material Description", yaxis_title="Months of Stock (NMOS)", height=400, xaxis_tickangle=-45)
                st.plotly_chart(fig_stock_out, use_container_width=True)
            else:
                st.success("✅ No stock out items found!")

            try:
                if 'Material Description' in df_filtered.columns and 'NMOS' in df_filtered.columns:
                    mos_cols_chart = ['Material Description', 'NMOS', 'GIT_MOS', 'LC_MOS', 'WB_MOS', 'TMD_MOS', 'TMOS']
                    available_cols = [c for c in mos_cols_chart if c in df_filtered.columns]
                    mos_df = df_filtered[available_cols].copy()
                    mos_df['NMOS'] = pd.to_numeric(mos_df['NMOS'], errors='coerce')
                    mos_df = mos_df[mos_df['NMOS'].notna()]
                    if not mos_df.empty:
                        for c in ['GIT_MOS', 'LC_MOS', 'WB_MOS', 'TMD_MOS', 'TMOS']:
                            if c in mos_df.columns:
                                mos_df[c] = pd.to_numeric(mos_df[c], errors='coerce').fillna(0)

                        mos_df = mos_df.sort_values('NMOS', ascending=True).reset_index(drop=True)
                        split_len = 40
                        mos_df['Material_split'] = mos_df['Material Description'].apply(lambda x: '<br>'.join([str(x)[i:i + split_len] for i in range(0, len(str(x)), split_len)]))
                        mos_df['NMOS_color'] = mos_df['NMOS'].apply(lambda x: "red" if x < 1 else "yellow" if x < 6 else "green" if x <= 18 else "skyblue")

                        split_size = 10
                        for i in range(0, len(mos_df), split_size):
                            df_chunk = mos_df.iloc[i:i + split_size].copy()
                            df_chunk = df_chunk.iloc[::-1].reset_index(drop=True)
                            fig = go.Figure()
                            fig.add_trace(go.Bar(y=df_chunk['Material_split'], x=df_chunk['NMOS'], name='NMOS', orientation='h', marker=dict(color=df_chunk['NMOS_color']), text=df_chunk['NMOS'].apply(lambda x: f"{x:.1f}" if pd.notna(x) else ""), textposition='inside', textfont_size=12, hovertemplate='NMOS: %{x:.1f} months<extra></extra>'))
                            pipeline_traces = [('GIT_MOS', 'cyan', 'GIT MOS'), ('LC_MOS', 'plum', 'LC MOS'), ('WB_MOS', 'gray', 'WB MOS'), ('TMD_MOS', 'orange', 'TMD MOS')]
                            for col, color, label in pipeline_traces:
                                if col in df_chunk.columns and (df_chunk[col] > 0).any():
                                    fig.add_trace(go.Bar(y=df_chunk['Material_split'], x=df_chunk[col], name=label, orientation='h', marker_color=color, text=df_chunk[col].apply(lambda x: f"{x:.1f}" if x > 0 else ""), textposition='inside', textfont_size=12, hovertemplate=f'{label}: %{{x:.1f}} months<extra></extra>'))
                            if 'TMOS' in df_chunk.columns:
                                fig.add_trace(go.Scatter(y=df_chunk['Material_split'], x=df_chunk['TMOS'], mode='text', text=df_chunk['TMOS'].apply(lambda x: f"TMOS: {x:.2f}" if x > 0 else ""), textposition='middle right', showlegend=False, textfont_size=12))
                            original_start = i + 1
                            original_end = i + len(df_chunk)
                            chart_title = f'{program_display} Stock Status - National and Pipeline (Medicines {original_start}-{original_end})'
                            fig.update_layout(barmode='stack', title=chart_title, xaxis_title='Months of Stock', yaxis_title='Material Description', height=max(500, 35 * len(df_chunk)), showlegend=True, legend=dict(orientation="v", yanchor="bottom", y=1.02, xanchor="right", x=1), margin=dict(l=20, r=120, t=60, b=20))
                            st.plotly_chart(fig, use_container_width=True)
            except Exception as e:
                st.error(f"Error creating MOS chart: {e}")

            try:
                if all(col in df_filtered.columns for col in ['Hubs', 'Head Office', 'NSOH', 'Material Description']):
                    hubs_vals = pd.to_numeric(df_filtered['Hubs'], errors='coerce').fillna(0)
                    ho_vals = pd.to_numeric(df_filtered['Head Office'], errors='coerce').fillna(0)
                    nsoh_vals = pd.to_numeric(df_filtered['NSOH'], errors='coerce')
                    valid_nsoh_mask = nsoh_vals.notna() & (nsoh_vals > 0)
                    if valid_nsoh_mask.any():
                        valid_indices = valid_nsoh_mask[valid_nsoh_mask].index
                        hubs_vals_valid = hubs_vals[valid_indices]
                        ho_vals_valid = ho_vals[valid_indices]
                        nsoh_vals_valid = nsoh_vals[valid_indices]
                        materials_valid = df_filtered.loc[valid_indices, 'Material Description']
                        hubs_pct = (hubs_vals_valid / nsoh_vals_valid * 100).fillna(0)
                        ho_pct = (ho_vals_valid / nsoh_vals_valid * 100).fillna(0)
                        nsoh_formatted = nsoh_vals_valid.apply(lambda x: f"{x:,.0f}")
                        bar_df = pd.DataFrame({'Material Description': materials_valid, 'Hubs%': hubs_pct, 'Head Office%': ho_pct, 'NSOH_display': nsoh_formatted}).reset_index(drop=True)
                        bar_df = bar_df.sort_values('Hubs%')
                        bar_df['Material_split'] = bar_df['Material Description'].apply(lambda x: '<br>'.join([str(x)[i:i + 25] for i in range(0, len(str(x)), 25)]))
                        n = 11
                        for i in range(0, len(bar_df), n):
                            df_chunk = bar_df.iloc[i:i + n]
                            fig_bar = go.Figure()
                            fig_bar.add_trace(go.Bar(y=df_chunk['Material_split'], x=df_chunk['Hubs%'], name='Hubs%', orientation='h', marker_color='skyblue', text=df_chunk['Hubs%'].apply(lambda x: f"{x:.1f}%" if x > 0 else ""), textposition='inside', textfont_size=12))
                            fig_bar.add_trace(go.Bar(y=df_chunk['Material_split'], x=df_chunk['Head Office%'], name='Head Office%', orientation='h', marker_color='orange', text=df_chunk['Head Office%'].apply(lambda x: f"{x:.1f}%" if x > 0 else ""), textposition='inside', textfont_size=12))
                            for idx, row in df_chunk.iterrows():
                                fig_bar.add_annotation(x=row['Hubs%'] + row['Head Office%'] + 2, y=row['Material_split'], text=f"NSOH: {row['NSOH_display']}", showarrow=False, font=dict(size=12), xanchor='left', yanchor='middle')
                            fig_bar.update_layout(barmode='stack', title=f'{program_display} Stock Distribution - Hubs vs Head Office (Materials {i + 1}-{i + len(df_chunk)})', xaxis_title='Percentage of NSOH (%)', yaxis_title='Material Description', xaxis={'range': [0, 120]}, height=max(600, 40 * len(df_chunk)), margin=dict(r=150))
                            st.plotly_chart(fig_bar, use_container_width=True)
                    else:
                        st.info("No materials with valid NSOH (>0) to display.")
            except Exception:
                pass
        else:
            st.info("No data available for KPI calculations.")

    # ---------------------------------------------------
    # TAB 3 - Decision Briefs (FIXED CARD VIEW)
    # ---------------------------------------------------
    with tab3:
        if sheet_name == "All":
            st.markdown("<h3 style='font-size: 28px; font-weight: bold; font-family: Times New Roman;'>All Programs - Medicines Needing Immediate Action</h3>", unsafe_allow_html=True)
        else:
            st.markdown(f"<h3 style='font-size: 28px; font-weight: bold; font-family: Times New Roman;'>{program_display} Medicines Needing Immediate Action</h3>", unsafe_allow_html=True)

        if not df_filtered.empty and 'Material Description' in df_filtered.columns:
            decision_df = df_filtered[['Material Description', 'NSOH', 'Expiry', 'AMC', 'NMOS', 'Status', 'Risk Type', 'Stock Status', 'Expiry Risk Details', 
                                       'GIT_MOS', 'LC_MOS', 'WB_MOS', 'TMD_MOS', 'GIT_PO', 'LC_PO', 'WB_PO', 'TMD_PO', 'Hubs%', 'Head Office%', 'CV Category']].copy()

            def get_identified_problems(row):
                if row.get('Stock Status') == 'Stock Out':
                    return "Stock Out"
                risk = row.get('Risk Type', '')
                if risk and risk != '':
                    return risk
                return ''

            decision_df['Identified Problems'] = decision_df.apply(get_identified_problems, axis=1)

            def get_automated_recommendation(row):
                problem = row.get('Identified Problems', '')
                if problem == 'Expiry Risk':
                    cv_cat = row.get('CV Category', 'Unknown')
                    return get_expiry_risk_recommendation(row, cv_cat)
                elif problem == 'Critical Risk':
                    stock_rec = get_stock_out_recommendation(row)
                    cv_cat = row.get('CV Category', 'Unknown')
                    expiry_rec = get_expiry_risk_recommendation(row, cv_cat)
                    return f"{stock_rec}\n{expiry_rec}"
                elif problem in ['Risk of Stock out', 'Stock Out']:
                    return get_stock_out_recommendation(row)
                return ''

            decision_df['Recommendation'] = decision_df.apply(get_automated_recommendation, axis=1)
            decision_df = decision_df[decision_df['Identified Problems'] != ''].copy()

            if len(decision_df) > 0:
                for col in ['NSOH', 'AMC']:
                    if col in decision_df.columns:
                        decision_df[col] = decision_df[col].apply(format_number_with_commas)
                if 'NMOS' in decision_df.columns:
                    decision_df['NMOS'] = decision_df['NMOS'].apply(format_mos_with_decimals)
                if 'Hubs%' in decision_df.columns:
                    decision_df['Hubs%'] = decision_df['Hubs%'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
                if 'Head Office%' in decision_df.columns:
                    decision_df['Head Office%'] = decision_df['Head Office%'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "")

                st.markdown("<h4 style='font-size: 24px; font-weight: bold; font-family: Times New Roman;'>Quick Summary</h4>", unsafe_allow_html=True)
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Items with Problems", len(decision_df))
                with col2:
                    st.metric("Stock Out Items", len(decision_df[decision_df['Identified Problems'] == 'Stock Out']))
                with col3:
                    st.metric("At Risk of Stock Out", len(decision_df[decision_df['Identified Problems'] == 'Risk of Stock out']))
                with col4:
                    st.metric("At Risk of Expiry", len(decision_df[decision_df['Identified Problems'] == 'Expiry Risk']) + len(decision_df[decision_df['Identified Problems'] == 'Critical Risk']))

                st.markdown("---")

                st.markdown("### ✏️ Sales and Operational Planning")
                st.info("💡 Recommendations are auto-generated based on pipeline status (with PO numbers) and distribution patterns. You can edit them as needed.")

                display_columns = ['Material Description', 'NSOH', 'Expiry', 'AMC', 'NMOS', 'Status', 'CV Category', 'Identified Problems', 'Recommendation']
                available_display_columns = [col for col in display_columns if col in decision_df.columns]

                column_config = {
                    "Material Description": st.column_config.TextColumn("Material", width=250, disabled=True, pinned=True),
                    "NSOH": st.column_config.TextColumn("NSOH", width=100, disabled=True),
                    "Expiry": st.column_config.TextColumn("Expiry", width=120, disabled=True),
                    "AMC": st.column_config.TextColumn("AMC", width=100, disabled=True),
                    "NMOS": st.column_config.TextColumn("NMOS", width=80, disabled=True),
                    "Status": st.column_config.TextColumn("Status", width=100, disabled=True),
                    "CV Category": st.column_config.TextColumn("CV Category", width=120, disabled=True, help="Coefficient of Variation category from hub distribution"),
                    "Identified Problems": st.column_config.TextColumn("Problem", width=120, disabled=True),
                    "Recommendation": st.column_config.TextColumn("Recommendation", width=450, disabled=False, help="Editable field - modify recommendation as needed")
                }

                edited_result = st.data_editor(
                    decision_df[available_display_columns],
                    column_config=column_config,
                    use_container_width=True, 
                    hide_index=True, 
                    height=min(600, (len(decision_df) + 1) * 45), 
                    num_rows="fixed"
                )

                for idx, row in edited_result.iterrows():
                    st.session_state.saved_recommendations[row['Material Description']] = {
                        'recommendation': row['Recommendation']
                    }

                st.download_button(
                    label="📥 Download Decision Briefs (CSV)", 
                    data=edited_result.to_csv(index=False), 
                    file_name=f"{sheet_name}_decision_briefs_{datetime.now().strftime('%Y%m%d')}.csv".replace(" ", "_"), 
                    mime="text/csv",
                    use_container_width=True
                )

                if st.button("🗑️ Clear All Recommendations", use_container_width=True):
                    st.session_state.saved_recommendations = {}
                    st.rerun()

                st.markdown("---")
                st.markdown("### 📇 Card View with Recommendations")

                for idx, row in decision_df.iterrows():
                    risk_type = row['Identified Problems']
                    if risk_type == "Critical Risk":
                        status_color = "#9b59b6"
                    elif risk_type == "Stock Out":
                        status_color = "#ff4444"
                    elif risk_type in ["Risk of Stock out", "Expiry Risk"]:
                        status_color = "#ffa500"
                    else:
                        status_color = "#ffa500"

                    recommendation_text = row.get('Recommendation', '')
                    cv_category = row.get('CV Category', 'N/A')

                    # FIXED: Safely get material description
                    material_desc = row.get('Material Description', 'N/A')
                    if pd.isna(material_desc) or material_desc is None:
                        material_desc = 'N/A'
                    else:
                        material_desc = str(material_desc)[:60]

                    st.markdown(f"""
                    <div class="stock-card" style='border-left: 4px solid {status_color}; margin-bottom: 15px;'>
                        <h4 style='color: {status_color}; margin-bottom: 10px;'>{material_desc}</h4>
                        <p><strong>📦 NSOH:</strong> {row.get('NSOH', 'N/A')}</p>
                        <p><strong>📅 Expiry:</strong> {row.get('Expiry', 'N/A')}</p>
                        <p><strong>📈 AMC:</strong> {row.get('AMC', 'N/A')}</p>
                        <p><strong>⏰ NMOS:</strong> <span style='color: {status_color}; font-weight: bold;'>{row.get('NMOS', 'N/A')}</span></p>
                        <p><strong>📊 Status:</strong> {row.get('Status', 'N/A')}</p>
                        <p><strong>⚠️ Risk Type:</strong> <span style='color: {status_color}; font-weight: bold;'>{risk_type}</span></p>
                        <p><strong>📊 CV Category:</strong> {cv_category}</p>
                        <p><strong>💡 Recommendation:</strong> <span style='background-color: #f0f2f6; padding: 5px; border-radius: 5px; display: inline-block;'>{recommendation_text}</span></p>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.success("✅ No medicines with identified problems! All items are within normal stock levels.")
                st.balloons()
        else:
            st.info("No data available.")

        # ---------------------------------------------------
    # TAB 4 - Hubs Distribution (WITH CORRECTED HEATMAP - EXCLUDING HUBS & HUBS%)
    # ---------------------------------------------------
    with tab4:
        try:
            if not df.empty:
                if branch_amc_data is not None and 'Material Description' in df.columns and 'Material Description' in branch_amc_data.columns and not branch_amc_data.empty:
                    branch_cols = [col for col in df.columns if 'Branch' in col or col == 'Material Description']
                    gh = df[branch_cols].copy() if branch_cols else pd.DataFrame()

                    st.markdown("<h4 style='font-size: 24px; font-weight: bold; font-family: Times New Roman;'>Stock Distribution Across Hubs by MOS</h4>", unsafe_allow_html=True)

                    merged_df = pd.merge(gh, branch_amc_data, on='Material Description', how='inner', suffixes=('_gh', '_cf'))

                    if not merged_df.empty:
                        gh_cols = [col for col in gh.columns if col != 'Material Description']
                        cf_cols = [col for col in branch_amc_data.columns if col != 'Material Description']
                        division_data = {'Material Description': merged_df['Material Description']}

                        min_cols = min(len(gh_cols), len(cf_cols))

                        for i in range(min_cols):
                            gh_col = gh_cols[i]
                            cf_col = cf_cols[i]
                            display_col_name = gh_col
                            gh_values = pd.to_numeric(merged_df[f"{gh_col}_gh"], errors='coerce')
                            cf_values = pd.to_numeric(merged_df[f"{cf_col}_cf"], errors='coerce')
                            with np.errstate(divide='ignore', invalid='ignore'):
                                division_result = np.where(cf_values != 0, gh_values / cf_values, np.nan)
                            division_data[display_col_name] = division_result

                        division_df = pd.DataFrame(division_data)
                        division_df = division_df.replace([np.inf, -np.inf], np.nan).round(2)

                        if division_df.shape[1] > 1:
                            branch_cols_list = [col for col in division_df.columns if col != 'Material Description']
                            division_df['CV (%)'] = division_df[branch_cols_list].apply(lambda row: calculate_coefficient_of_variation(row), axis=1)
                            division_df['CV (%)'] = division_df['CV (%)'].round(1)
                            cols = ['Material Description', 'CV (%)'] + branch_cols_list
                            division_df = division_df[cols]

                            def categorize_cv(cv_value):
                                if pd.isna(cv_value):
                                    return "Unknown"
                                elif cv_value < 50:
                                    return "Low variation"
                                elif cv_value <= 100:
                                    return "Moderate variation"
                                else:
                                    return "High variation"

                            division_df['CV Category'] = division_df['CV (%)'].apply(categorize_cv)
                            cv_counts = division_df['CV Category'].value_counts()

                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("Total Materials", len(division_df))
                            if 'Low variation' in cv_counts:
                                with col2:
                                    low_count = cv_counts['Low variation']
                                    low_pct = (low_count / len(division_df) * 100) if len(division_df) > 0 else 0
                                    st.metric("Low Variation (<50%)", f"{low_count} ({low_pct:.1f}%)")
                            if 'Moderate variation' in cv_counts:
                                with col3:
                                    mod_count = cv_counts['Moderate variation']
                                    mod_pct = (mod_count / len(division_df) * 100) if len(division_df) > 0 else 0
                                    st.metric("Moderate Variation (50-100%)", f"{mod_count} ({mod_pct:.1f}%)")
                            if 'High variation' in cv_counts:
                                with col4:
                                    high_count = cv_counts['High variation']
                                    high_pct = (high_count / len(division_df) * 100) if len(division_df) > 0 else 0
                                    st.metric("High Variation (>100%)", f"{high_count} ({high_pct:.1f}%)")

                            # MOS Heatmap (existing)
                            if division_df.shape[1] > 2:
                                branch_cols_list = [col for col in division_df.columns if col not in ['Material Description', 'CV (%)', 'CV Category']]
                                heatmap_df = division_df.copy()
                                heatmap_df = heatmap_df.sort_values('Material Description')
                                heatmap_df_indexed = heatmap_df.set_index('Material Description')
                                heatmap_df_indexed = heatmap_df_indexed[branch_cols_list]
                                heatmap_df_transposed = heatmap_df_indexed.T
                                total_materials = len(heatmap_df_transposed.columns)
                                materials_per_page = 11

                                if total_materials > materials_per_page:
                                    total_pages = (total_materials + materials_per_page - 1) // materials_per_page
                                    col1, col2, col3 = st.columns([1, 3, 1])
                                    with col1:
                                        if st.button("◀ Previous", key="heatmap_prev"):
                                            if st.session_state.heatmap_page > 1:
                                                st.session_state.heatmap_page -= 1
                                                st.rerun()
                                    with col2:
                                        st.markdown(f"<h5 style='text-align: center;'>Page {st.session_state.heatmap_page} of {total_pages}</h5>", unsafe_allow_html=True)
                                    with col3:
                                        if st.button("Next ▶", key="heatmap_next"):
                                            if st.session_state.heatmap_page < total_pages:
                                                st.session_state.heatmap_page += 1
                                                st.rerun()
                                    start_idx = (st.session_state.heatmap_page - 1) * materials_per_page
                                    end_idx = min(start_idx + materials_per_page, total_materials)
                                    page_materials = heatmap_df_transposed.columns[start_idx:end_idx]
                                    heatmap_page_df = heatmap_df_transposed[page_materials]
                                    st.info(f"Showing materials {start_idx + 1} to {end_idx} of {total_materials}")
                                else:
                                    heatmap_page_df = heatmap_df_transposed

                                fig = go.Figure(data=go.Heatmap(
                                    z=heatmap_page_df.values, y=heatmap_page_df.index, x=heatmap_page_df.columns,
                                    colorscale=[[0.0, 'red'], [0.125, 'yellow'], [0.5, 'green'], [1.0, 'skyblue']],
                                    zmin=0, zmax=8, text=heatmap_page_df.values.round(1), texttemplate='%{text}', textfont={"size": 14},
                                    colorbar=dict(title="MOS", tickvals=[0.5, 1, 2, 4, 6, 8], ticktext=['0.5', '1', '2', '4', '6', '8+']),
                                    hovertemplate='<b>Material:</b> %{x}<br><b>Branch:</b> %{y}<br><b>MOS:</b> %{z:.2f} months<br><extra></extra>'
                                ))
                                fig.update_layout(xaxis={'title': 'Material Description', 'tickangle': -45, 'tickfont': {'size': 12}}, 
                                                yaxis={'title': 'Branches', 'tickfont': {'size': 12}}, height=650, margin=dict(l=120, r=120, t=50, b=200))
                                st.plotly_chart(fig, use_container_width=True)

                                st.markdown("#### MOS Thresholds:")
                                col1, col2, col3, col4 = st.columns(4)
                                with col1:
                                    st.markdown("🔴 **< 0.5** : Stock Out")
                                with col2:
                                    st.markdown("🟡 **0.5 - 1** : Understock")
                                with col3:
                                    st.markdown("🟢 **1 - 4** : Normal Stock")
                                with col4:
                                    st.markdown("🔵 **> 4** : Overstock")

                                # ========== NEW: EPSS Hubs SOH Heatmap (Excluding "Hubs" and "Hubs%") ==========
                                st.markdown("---")
                                st.markdown("<h5 style='font-size: 20px; font-weight: bold;'>🔥 EPSS Hubs Stock on Hand (SOH) Heatmap</h5>", unsafe_allow_html=True)
                                st.caption("🔴 Red shades indicate higher stock levels | Lighter shades indicate lower stock levels")

                                # Get branch stock columns for heatmap - EXCLUDE "Hubs" and "Hubs%"
                                branch_soh_cols = []
                                for col in df.columns:
                                    # Include columns with 'Branch' in name
                                    if 'Branch' in col:
                                        branch_soh_cols.append(col)
                                    # Include columns that start with 'Hub' but NOT exactly 'Hubs' or 'Hubs%'
                                    elif col.startswith('Hub') and col not in ['Hubs', 'Hubs%']:
                                        branch_soh_cols.append(col)
                                    # Also include any column that has specific hub names (Adama, Addis, etc.)
                                    elif col in ['Adama', 'Addis Ababa', 'Bahir Dar', 'Gondar', 'Mekele', 'Hawassa', 'Jimma', 'Dire Dawa', 'Dessie', 'Arba Minch', 'Jigjiga', 'Assosa', 'Gambela', 'Semera', 'Shire', 'Kebridahar', 'Negele Borena', 'Nekemte']:
                                        branch_soh_cols.append(col)

                                # Remove duplicates and Material Description
                                branch_soh_cols = list(set(branch_soh_cols))
                                branch_soh_cols = [col for col in branch_soh_cols if col != 'Material Description']

                                if branch_soh_cols:
                                    # Prepare data for heatmap - Materials on X-axis, Branches on Y-axis
                                    heatmap_data = df[['Material Description'] + branch_soh_cols].copy()

                                    # Convert to numeric and fill NaN with 0
                                    for col in branch_soh_cols:
                                        heatmap_data[col] = pd.to_numeric(heatmap_data[col], errors='coerce').fillna(0)

                                    # Set Material Description as index (rows = materials)
                                    heatmap_data = heatmap_data.set_index('Material Description')

                                    # Transpose so that Materials are on X-axis, Branches on Y-axis
                                    heatmap_data_transposed = heatmap_data.T  # Branches become rows, Materials become columns

                                    # Pagination: Split materials (columns) into pages of 10
                                    total_materials_soh = len(heatmap_data_transposed.columns)
                                    materials_per_page_soh = 10  # Show 10 materials per page

                                    if 'soh_heatmap_page' not in st.session_state:
                                        st.session_state.soh_heatmap_page = 1

                                    if total_materials_soh > materials_per_page_soh:
                                        total_pages_soh = (total_materials_soh + materials_per_page_soh - 1) // materials_per_page_soh

                                        col_prev, col_page, col_next = st.columns([1, 3, 1])
                                        with col_prev:
                                            if st.button("◀ Previous", key="soh_heatmap_prev"):
                                                if st.session_state.soh_heatmap_page > 1:
                                                    st.session_state.soh_heatmap_page -= 1
                                                    st.rerun()
                                        with col_page:
                                            st.markdown(f"<p style='text-align: center;'>Page {st.session_state.soh_heatmap_page} of {total_pages_soh}</p>", unsafe_allow_html=True)
                                        with col_next:
                                            if st.button("Next ▶", key="soh_heatmap_next"):
                                                if st.session_state.soh_heatmap_page < total_pages_soh:
                                                    st.session_state.soh_heatmap_page += 1
                                                    st.rerun()

                                        start_idx_soh = (st.session_state.soh_heatmap_page - 1) * materials_per_page_soh
                                        end_idx_soh = min(start_idx_soh + materials_per_page_soh, total_materials_soh)
                                        # Select a subset of columns (materials) for current page
                                        page_materials_soh = heatmap_data_transposed.columns[start_idx_soh:end_idx_soh]
                                        heatmap_page_df_soh = heatmap_data_transposed[page_materials_soh]
                                        st.info(f"Showing materials {start_idx_soh + 1} to {end_idx_soh} of {total_materials_soh}")
                                    else:
                                        heatmap_page_df_soh = heatmap_data_transposed

                                    if not heatmap_page_df_soh.empty:
                                        # Create heatmap with Reds color scale
                                        # X-axis: Materials, Y-axis: Branches
                                        fig_soh = go.Figure(data=go.Heatmap(
                                            z=heatmap_page_df_soh.values,  # Stock values
                                            y=heatmap_page_df_soh.index,   # Branches (Y-axis)
                                            x=heatmap_page_df_soh.columns, # Materials (X-axis)
                                            colorscale='Reds',
                                            zmin=0,
                                            zmax=heatmap_page_df_soh.max().max(),
                                            text=heatmap_page_df_soh.values,
                                            texttemplate='%{text:,.0f}',
                                            textfont={"size": 12},
                                            colorbar=dict(
                                                title="Stock on Hand (Units)",
                                                title_side="right",
                                                ticks="outside"
                                            ),
                                            hovertemplate='<b>Material:</b> %{x}<br>' +
                                                          '<b>Branch:</b> %{y}<br>' +
                                                          '<b>SOH:</b> %{z:,.0f} units<br>' +
                                                          '<extra></extra>'
                                        ))

                                        fig_soh.update_layout(
                                            title="EPSS Hubs Stock on Hand Distribution",
                                            xaxis={
                                                'title': 'Material Description',
                                                'tickangle': -45,
                                                'tickfont': {'size': 10},
                                                'automargin': True
                                            },
                                            yaxis={
                                                'title': 'Branches / Hubs',
                                                'tickfont': {'size': 11}
                                            },
                                            height=max(500, 35 * len(heatmap_page_df_soh.index)),
                                            margin=dict(l=150, r=50, t=80, b=200)
                                        )

                                        st.plotly_chart(fig_soh, use_container_width=True)

                                        # Heatmap summary statistics
                                        st.markdown("**Summary Statistics for Current Page:**")
                                        col_h1, col_h2, col_h3, col_h4 = st.columns(4)
                                        with col_h1:
                                            page_total = heatmap_page_df_soh.sum().sum()
                                            st.metric("📦 Total Stock (Page)", f"{int(page_total):,}")
                                        with col_h2:
                                            max_branch = heatmap_page_df_soh.sum(axis=1).idxmax() if not heatmap_page_df_soh.empty else "N/A"
                                            max_branch_value = int(heatmap_page_df_soh.sum(axis=1).max()) if not heatmap_page_df_soh.empty else 0
                                            st.metric("🏆 Hub with Most Stock", max_branch, delta=f"{max_branch_value:,} units")
                                        with col_h3:
                                            max_material = heatmap_page_df_soh.sum(axis=0).idxmax() if not heatmap_page_df_soh.empty else "N/A"
                                            max_material_value = int(heatmap_page_df_soh.sum(axis=0).max()) if not heatmap_page_df_soh.empty else 0
                                            st.metric("🥇 Material with Most Stock", max_material[:25], delta=f"{max_material_value:,} units")
                                        with col_h4:
                                            # Average stock per material on this page
                                            avg_stock = heatmap_page_df_soh.mean(axis=0).mean() if not heatmap_page_df_soh.empty else 0
                                            st.metric("📊 Avg Stock per Material", f"{int(avg_stock):,}")

                                        # Overall summary (all materials)
                                        with st.expander("📊 View Overall Summary (All Materials)"):
                                            total_stock_all = heatmap_data.sum().sum()
                                            hub_max_all = heatmap_data.T.sum(axis=1).idxmax()
                                            hub_max_value_all = int(heatmap_data.T.sum(axis=1).max())
                                            material_max_all = heatmap_data.sum(axis=1).idxmax()
                                            material_max_value_all = int(heatmap_data.sum(axis=1).max())

                                            col_a, col_b, col_c = st.columns(3)
                                            with col_a:
                                                st.metric("📦 Total Stock All Hubs", f"{int(total_stock_all):,}")
                                            with col_b:
                                                st.metric("🏆 Hub with Most Stock (Overall)", hub_max_all, delta=f"{hub_max_value_all:,} units")
                                            with col_c:
                                                st.metric("🥇 Material with Most Stock (Overall)", material_max_all[:35], delta=f"{material_max_value_all:,} units")
                                    else:
                                        st.info("No data available for heatmap")
                                else:
                                    st.info("No branch/hub stock data available for heatmap")

                                st.markdown("---")
                                st.markdown("<h5 style='font-size: 20px; font-weight: bold;'>Full Branch MOS Data with CV</h5>", unsafe_allow_html=True)
                                display_division_df = division_df.copy()
                                display_division_df['CV (%)'] = display_division_df['CV (%)'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
                                st.dataframe(display_division_df, use_container_width=True, height=400, hide_index=True)
                                st.download_button(label="Download Hubs MOS with CV", data=division_df.to_csv(index=False), file_name="hubs_mos_distribution_with_cv.csv", mime="text/csv")

                                st.markdown("---")
                                st.markdown("<h5 style='font-size: 20px; font-weight: bold;'>EPSS Hubs SOH</h5>", unsafe_allow_html=True)
                                st.dataframe(gh, use_container_width=True, height=400, hide_index=True)

                                st.markdown("---")
                                st.markdown("<h5 style='font-size: 20px; font-weight: bold;'>EPSS Hubs AMC Data (from Google Sheets)</h5>", unsafe_allow_html=True)
                                st.dataframe(branch_amc_data, use_container_width=True, height=400, hide_index=True)
                            else:
                                st.warning("Not enough branch columns for heatmap")
                        else:
                            st.warning("No matching Material Description found")
                    elif branch_amc_data is None or branch_amc_data.empty:
                        st.info("Branch AMC data is currently unavailable.")
                        if not df.empty and 'Material Description' in df.columns:
                            branch_cols = [col for col in df.columns if 'Branch' in col or col == 'Material Description']
                            gh = df[branch_cols].copy() if branch_cols else df[['Material Description']].copy()
                            st.dataframe(gh, use_container_width=True, height=400, hide_index=True)
                    else:
                        st.error("'Material Description' column not found")
                else:
                    st.warning("Main dataframe is empty.")
            else:
                st.warning("Main dataframe is empty.")
        except Exception as e:
            st.error(f"Error processing files: {e}")

    # ---------------------------------------------------
    # TAB 5 - Supply Planning
    # ---------------------------------------------------
    with tab5:
        st.markdown("<h3 style='font-size: 24px; font-weight: bold;'>📦 Supply Planning - Procurement Requirements</h3>", unsafe_allow_html=True)

        with st.expander("📖 Parameters & Instructions", expanded=False):
            st.markdown("""
            **Supply Planning Parameters:**
            - Lead Time = 6 months (time from order placement to delivery)
            - Safety Stock = 2 months (buffer stock)
            - Maximum Stock Level = 18 months
            - Reorder Point = Lead Time + Safety Stock = 8 months

            **Order Quantity Formula:**
            - Order Quantity = (18 - TMOS) × AMC (ONLY if 18 - TMOS is POSITIVE)
            - MOS Needed = 18 - TMOS (months of stock required to reach maximum)

            **TMOS = NMOS + Pipeline MOS** (GIT_MOS + LC_MOS + WB_MOS + TMD_MOS)
            """)

        if 'TMOS' in df_filtered.columns and 'AMC' in df_filtered.columns and 'NMOS' in df_filtered.columns:
            supply_plan = []
            current_date = datetime.now()

            def get_future_date(months_from_now):
                target_date = current_date
                months_int = int(months_from_now)
                new_year = target_date.year + ((target_date.month + months_int - 1) // 12)
                new_month = ((target_date.month + months_int - 1) % 12) + 1
                target_date = target_date.replace(year=new_year, month=new_month, day=1)
                return target_date

            def get_readable_order_by(months_until_order):
                if months_until_order <= 0:
                    return "Now"
                target_date = get_future_date(months_until_order)
                month_name = target_date.strftime('%B')
                year = target_date.year
                if target_date.day <= 15:
                    period = "beginning"
                else:
                    period = "end"
                if target_date.year == current_date.year:
                    return f"{period} of {month_name}"
                else:
                    return f"{period} of {month_name} {year}"

            for idx, row in df_filtered.iterrows():
                tmos = row.get('TMOS', 0)
                amc = row.get('AMC', 0)
                nmos = row.get('NMOS', 0)
                material = row['Material Description']

                git_mos = row.get('GIT_MOS', 0)
                lc_mos = row.get('LC_MOS', 0)
                wb_mos = row.get('WB_MOS', 0)
                tmd_mos = row.get('TMD_MOS', 0)

                try:
                    tmos = float(tmos) if pd.notna(tmos) else 0
                    amc = float(amc) if pd.notna(amc) else 0
                    nmos = float(nmos) if pd.notna(nmos) else 0
                    git_mos = float(git_mos) if pd.notna(git_mos) else 0
                    lc_mos = float(lc_mos) if pd.notna(lc_mos) else 0
                    wb_mos = float(wb_mos) if pd.notna(wb_mos) else 0
                    tmd_mos = float(tmd_mos) if pd.notna(tmd_mos) else 0
                except:
                    continue

                pipeline_mos = git_mos + lc_mos + wb_mos + tmd_mos
                mos_needed = 18 - tmos

                if mos_needed > 0 and amc > 0:
                    order_quantity = int(mos_needed * amc)

                    if tmos <= 8:
                        urgency = "🔴 CRITICAL"
                        action = f"Place this {order_quantity:,} units IMMEDIATELY"
                        order_by = "Now"
                        expected_delivery = get_future_date(6)
                    else:
                        months_until_order = round(tmos - 8, 1)
                        urgency = "🟡 PLAN"
                        order_by_readable = get_readable_order_by(months_until_order)
                        action = f"Place this {order_quantity:,} units by {order_by_readable}"
                        order_by = order_by_readable
                        total_months_to_delivery = months_until_order + 6
                        expected_delivery = get_future_date(total_months_to_delivery)

                    pipeline_parts = []
                    if git_mos > 0:
                        pipeline_parts.append(f"GIT: {round(git_mos,1)}m")
                    if lc_mos > 0:
                        pipeline_parts.append(f"LC: {round(lc_mos,1)}m")
                    if wb_mos > 0:
                        pipeline_parts.append(f"WB: {round(wb_mos,1)}m")
                    if tmd_mos > 0:
                        pipeline_parts.append(f"TMD: {round(tmd_mos,1)}m")
                    pipeline_status = ", ".join(pipeline_parts) if pipeline_parts else "No pipeline stock"

                    supply_plan.append({
                        'Material': material,
                        'Current TMOS': round(tmos, 2),
                        'NMOS': round(nmos, 2),
                        'Pipeline': round(pipeline_mos, 2),
                        'Pipeline Status': pipeline_status,
                        'AMC': int(amc),
                        'MOS Needed': round(mos_needed, 2),
                        'Order Quantity': f"{order_quantity:,}",
                        'Urgency': urgency,
                        'Action': action,
                        'Order By': order_by,
                        'Expected Delivery': expected_delivery.strftime('%b %Y')
                    })

            if supply_plan:
                supply_df = pd.DataFrame(supply_plan).sort_values('Current TMOS', ascending=True)

                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric("📋 Materials to Order", len(supply_df))
                with col2:
                    critical = len([s for s in supply_plan if s['Urgency'] == '🔴 CRITICAL'])
                    st.metric("🔴 Critical", critical, delta="Order Now")
                with col3:
                    plan = len([s for s in supply_plan if s['Urgency'] == '🟡 PLAN'])
                    st.metric("🟡 Plan", plan, delta="Future Order")
                with col4:
                    total_quantity = sum([int(s['Order Quantity'].replace(',', '')) for s in supply_plan])
                    st.metric("📦 Total Order Qty", f"{total_quantity:,} units")
                with col5:
                    avg_mos_needed = supply_df['MOS Needed'].mean()
                    st.metric("📊 Avg MOS Needed", f"{round(avg_mos_needed, 1)} months")

                st.dataframe(
                    supply_df[['Material', 'Current TMOS', 'NMOS', 'Pipeline', 'AMC', 'MOS Needed', 'Order Quantity', 'Urgency', 'Action', 'Order By', 'Expected Delivery']],
                    use_container_width=True,
                    hide_index=True
                )

                st.download_button(
                    label="📥 Download Order Quantity Plan (CSV)",
                    data=supply_df.to_csv(index=False),
                    file_name=f"order_quantity_plan_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )

                # Action Plan Table
                st.markdown("---")
                st.markdown("<h4 style='font-size: 20px; font-weight: bold;'>📝 Action Plan - Materials Requiring Attention</h4>", unsafe_allow_html=True)

                # Initialize selected_tab in session state if not exists
                if 'action_plan_tab' not in st.session_state:
                    st.session_state.action_plan_tab = "📋 All Issues"

                selected_tab_action = st.session_state.action_plan_tab

                # Stylish filter buttons
                col_filter1, col_filter2, col_filter3, col_filter4, col_filter5, col_filter6 = st.columns(6)

                with col_filter1:
                    if st.button("📋 All Issues", use_container_width=True, type="primary" if selected_tab_action == "📋 All Issues" else "secondary"):
                        st.session_state.action_plan_tab = "📋 All Issues"
                        st.rerun()
                with col_filter2:
                    if st.button("🔴 Stock Out", use_container_width=True, type="primary" if selected_tab_action == "🔴 Stock Out" else "secondary"):
                        st.session_state.action_plan_tab = "🔴 Stock Out"
                        st.rerun()
                with col_filter3:
                    if st.button("🟡 Risk of SO", use_container_width=True, type="primary" if selected_tab_action == "🟡 Risk of Stock Out" else "secondary"):
                        st.session_state.action_plan_tab = "🟡 Risk of Stock Out"
                        st.rerun()
                with col_filter4:
                    if st.button("⚠️ Expiry Risk", use_container_width=True, type="primary" if selected_tab_action == "⚠️ Expiry Risk" else "secondary"):
                        st.session_state.action_plan_tab = "⚠️ Expiry Risk"
                        st.rerun()
                with col_filter5:
                    if st.button("📉 Below Min", use_container_width=True, type="primary" if selected_tab_action == "📉 Below Min Stock" else "secondary"):
                        st.session_state.action_plan_tab = "📉 Below Min Stock"
                        st.rerun()
                with col_filter6:
                    if st.button("📦 Pipeline Insuff", use_container_width=True, type="primary" if selected_tab_action == "📦 Pipeline Insufficient" else "secondary"):
                        st.session_state.action_plan_tab = "📦 Pipeline Insufficient"
                        st.rerun()

                st.markdown("---")

                action_plan = []

                for idx, row in df_filtered.iterrows():
                    material = row['Material Description']
                    nmos = row.get('NMOS', 0)
                    tmos = row.get('TMOS', 0)
                    nsoh = row.get('NSOH', 0)
                    amc = row.get('AMC', 0)
                    stock_status = row.get('Stock Status', '')
                    risk_type = row.get('Risk Type', '')
                    has_expiry_risk = row.get('Has Expiry Risk', False)
                    risk_of_stock = row.get('Risk of Stock', '')

                    git_mos = row.get('GIT_MOS', 0)
                    lc_mos = row.get('LC_MOS', 0)
                    wb_mos = row.get('WB_MOS', 0)
                    tmd_mos = row.get('TMD_MOS', 0)

                    git_po = row.get('GIT_PO', '')
                    lc_po = row.get('LC_PO', '')
                    wb_po = row.get('WB_PO', '')
                    tmd_po = row.get('TMD_PO', '')

                    try:
                        nmos = float(nmos) if pd.notna(nmos) else 0
                        tmos = float(tmos) if pd.notna(tmos) else 0
                        nsoh = float(nsoh) if pd.notna(nsoh) else 0
                        amc = float(amc) if pd.notna(amc) else 0
                        git_mos = float(git_mos) if pd.notna(git_mos) else 0
                        lc_mos = float(lc_mos) if pd.notna(lc_mos) else 0
                        wb_mos = float(wb_mos) if pd.notna(wb_mos) else 0
                        tmd_mos = float(tmd_mos) if pd.notna(tmd_mos) else 0
                    except:
                        continue

                    if amc == 0 and not has_expiry_risk:
                        continue

                    pmos = git_mos + lc_mos + wb_mos + tmd_mos
                    problems_list = []

                    current_date = datetime.now()
                    current_month = current_date.month
                    current_year = current_date.year

                    def get_end_of_month_date(year, month):
                        year_int = int(year)
                        month_int = int(month)
                        if month_int == 12:
                            next_month = 1
                            next_year = year_int + 1
                        else:
                            next_month = month_int + 1
                            next_year = year_int
                        last_day = (datetime(next_year, next_month, 1) - timedelta(days=1)).day
                        return datetime(year_int, month_int, last_day)

                    def has_pipeline():
                        return pmos > 0

                    def get_pipeline_recommendation():
                        recommendations = []
                        responsible = []

                        if git_mos > 0 and git_po != '' and str(git_po) != 'nan':
                            recommendations.append(f"🚚 Expedite shipment for GIT PO: {git_po}")
                            responsible.append("EPSS_CMD")
                        if lc_mos > 0 and lc_po != '' and str(lc_po) != 'nan':
                            recommendations.append(f"📄 Expedite L/C opening for PO: {lc_po}")
                            responsible.append("EPSS_CMD, EPSS_DMD")
                        if wb_mos > 0 and wb_po != '' and str(wb_po) != 'nan':
                            recommendations.append(f"💰 Expedite budget transfer for PO: {wb_po}")
                            responsible.append("EPSS_Finance, MOH")
                        if tmd_mos > 0 and tmd_po != '' and str(tmd_po) != 'nan':
                            recommendations.append(f"📋 Expedite tender process for PO: {tmd_po}")
                            responsible.append("EPSS_PMD, EPSS_DMD")

                        if recommendations:
                            return " | ".join(recommendations), ", ".join(set(responsible))
                        return "⚠️ Pipeline exists but no PO details available", "EPSS_DMD"

                    # Check for STOCK OUT
                    if stock_status == 'Stock Out' or nmos < 1:
                        if has_pipeline():
                            action_point, responsible_body = get_pipeline_recommendation()
                        else:
                            order_qty = int((18 - tmos) * amc) if amc > 0 else 0
                            action_point = f"📦 Place urgent order for {order_qty:,} units - no pipeline stock"
                            responsible_body = "EPSS_CMD, EPSS_DMD"

                        end_of_month = get_end_of_month_date(current_year, current_month)
                        due_date = end_of_month.strftime('Before %d %b %Y')

                        problems_list.append({
                            'problem': '🔴 Stock Out',
                            'action_point': action_point,
                            'responsible_body': responsible_body,
                            'due_date': due_date
                        })

                    # Check for RISK OF STOCK OUT
                    if (risk_of_stock == 'Risk of Stock out' or risk_type == 'Risk of Stock out') and nmos < 6 and nmos >= 1:
                        if has_pipeline():
                            action_point, responsible_body = get_pipeline_recommendation()
                        else:
                            order_qty = int((18 - tmos) * amc) if amc > 0 else 0
                            action_point = f"📦 Place order for {order_qty:,} units - initiate procurement (no pipeline stock)"
                            responsible_body = "EPSS_DMD"

                        end_of_month = get_end_of_month_date(current_year, current_month)
                        due_date = end_of_month.strftime('Before %d %b %Y')

                        problems_list.append({
                            'problem': '🟡 Risk of Stock Out',
                            'action_point': action_point,
                            'responsible_body': responsible_body,
                            'due_date': due_date
                        })

                    # Check for EXPIRY RISK
                    if has_expiry_risk or risk_type == 'Expiry Risk':
                        action_point = get_expiry_risk_recommendation(row)
                        responsible_body = "EPSS_DMD"
                        due_date = "ASAP"

                        problems_list.append({
                            'problem': '⚠️ Expiry Risk',
                            'action_point': action_point,
                            'responsible_body': responsible_body,
                            'due_date': due_date
                        })

                    # Check for BELOW MINIMUM STOCK LEVEL
                    if nmos < 6 and nmos >= 1 and not has_expiry_risk and not (risk_of_stock == 'Risk of Stock out' or risk_type == 'Risk of Stock out'):
                        if has_pipeline():
                            action_point, responsible_body = get_pipeline_recommendation()
                        else:
                            order_qty = int((6 - nmos) * amc) if amc > 0 else 0
                            action_point = f"📦 Place order for {order_qty:,} units to reach minimum stock level (6 months) - no pipeline stock"
                            responsible_body = "EPSS_DMD"

                        end_of_month = get_end_of_month_date(current_year, current_month)
                        due_date = end_of_month.strftime('Before %d %b %Y')

                        problems_list.append({
                            'problem': '📉 Below Minimum Stock Level',
                            'action_point': action_point,
                            'responsible_body': responsible_body,
                            'due_date': due_date
                        })

                    # Check for PIPELINE INSUFFICIENT
                    if tmos < 18 and nmos >= 6 and not has_expiry_risk:
                        if has_pipeline():
                            action_point, responsible_body = get_pipeline_recommendation()
                        else:
                            order_qty = int((18 - tmos) * amc) if amc > 0 else 0
                            action_point = f"📦 Place order for {order_qty:,} units to reach maximum stock level (18 months) - no pipeline stock"
                            responsible_body = "MOH"

                        due_date = "Plan for next quarter"

                        problems_list.append({
                            'problem': '📦 Pipeline Insufficient - Cannot Reach Max Stock',
                            'action_point': action_point,
                            'responsible_body': responsible_body,
                            'due_date': due_date
                        })

                    if not problems_list:
                        continue

                    nsoh_formatted = f"{int(nsoh):,}" if nsoh > 0 else "0"
                    amc_formatted = f"{int(amc):,}" if amc > 0 else "N/A"
                    pmos_formatted = f"{round(pmos, 2)}" if pmos > 0 else "0"

                    for problem_item in problems_list:
                        action_plan.append({
                            'Material': material,
                            'NSOH': nsoh_formatted,
                            'AMC': amc_formatted,
                            'PMOS': pmos_formatted,
                            'NMOS': round(nmos, 2),
                            'TMOS': round(tmos, 2),
                            'Identified Problem': problem_item['problem'],
                            'Action Point': problem_item['action_point'],
                            'Responsible Body': problem_item['responsible_body'],
                            'Due Date': problem_item['due_date']
                        })

                if action_plan:
                    action_df = pd.DataFrame(action_plan)

                    total_items = len(action_df)
                    stock_out_count = len([a for a in action_plan if a['Identified Problem'] == '🔴 Stock Out'])
                    risk_count = len([a for a in action_plan if a['Identified Problem'] == '🟡 Risk of Stock Out'])
                    expiry_count = len([a for a in action_plan if a['Identified Problem'] == '⚠️ Expiry Risk'])
                    below_min_count = len([a for a in action_plan if a['Identified Problem'] == '📉 Below Minimum Stock Level'])
                    pipeline_insufficient_count = len([a for a in action_plan if a['Identified Problem'] == '📦 Pipeline Insufficient - Cannot Reach Max Stock'])

                    # Metric cards
                    st.markdown("""
                    <style>
                    .metric-card {
                        background: white;
                        border-radius: 15px;
                        padding: 15px;
                        text-align: center;
                        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                        transition: transform 0.3s ease;
                    }
                    .metric-card:hover {
                        transform: translateY(-5px);
                        box-shadow: 0 8px 15px rgba(0,0,0,0.2);
                    }
                    .metric-value {
                        font-size: 28px;
                        font-weight: bold;
                        margin: 10px 0;
                    }
                    .metric-label {
                        font-size: 14px;
                        color: #666;
                    }
                    </style>
                    """, unsafe_allow_html=True)

                    col1, col2, col3, col4, col5, col6 = st.columns(6)

                    with col1:
                        st.markdown(f"""
                        <div class="metric-card">
                            <div>📋</div>
                            <div class="metric-value">{total_items}</div>
                            <div class="metric-label">Total Issues</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col2:
                        st.markdown(f"""
                        <div class="metric-card">
                            <div>🔴</div>
                            <div class="metric-value" style="color:#ff4444;">{stock_out_count}</div>
                            <div class="metric-label">Stock Out</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col3:
                        st.markdown(f"""
                        <div class="metric-card">
                            <div>🟡</div>
                            <div class="metric-value" style="color:#ffa500;">{risk_count}</div>
                            <div class="metric-label">Risk of SO</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col4:
                        st.markdown(f"""
                        <div class="metric-card">
                            <div>⚠️</div>
                            <div class="metric-value" style="color:#ff9800;">{expiry_count}</div>
                            <div class="metric-label">Expiry Risk</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col5:
                        st.markdown(f"""
                        <div class="metric-card">
                            <div>📉</div>
                            <div class="metric-value" style="color:#2196F3;">{below_min_count}</div>
                            <div class="metric-label">Below Min</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col6:
                        st.markdown(f"""
                        <div class="metric-card">
                            <div>📦</div>
                            <div class="metric-value" style="color:#9C27B0;">{pipeline_insufficient_count}</div>
                            <div class="metric-label">Pipeline Insuff</div>
                        </div>
                        """, unsafe_allow_html=True)

                    st.markdown("---")

                    # APPLY FILTER
                    if selected_tab_action == "🔴 Stock Out":
                        filtered_df = action_df[action_df['Identified Problem'] == '🔴 Stock Out']
                        st.info(f"📌 Showing {len(filtered_df)} items with STOCK OUT")
                    elif selected_tab_action == "🟡 Risk of Stock Out":
                        filtered_df = action_df[action_df['Identified Problem'] == '🟡 Risk of Stock Out']
                        st.info(f"📌 Showing {len(filtered_df)} items with RISK OF STOCK OUT")
                    elif selected_tab_action == "⚠️ Expiry Risk":
                        filtered_df = action_df[action_df['Identified Problem'] == '⚠️ Expiry Risk']
                        st.info(f"📌 Showing {len(filtered_df)} items with EXPIRY RISK")
                    elif selected_tab_action == "📉 Below Min Stock":
                        filtered_df = action_df[action_df['Identified Problem'] == '📉 Below Minimum Stock Level']
                        st.info(f"📌 Showing {len(filtered_df)} items with BELOW MINIMUM STOCK LEVEL")
                    elif selected_tab_action == "📦 Pipeline Insufficient":
                        filtered_df = action_df[action_df['Identified Problem'] == '📦 Pipeline Insufficient - Cannot Reach Max Stock']
                        st.info(f"📌 Showing {len(filtered_df)} items with PIPELINE INSUFFICIENT")
                    else:
                        filtered_df = action_df
                        st.info(f"📌 Showing all {len(filtered_df)} action items")

                    st.dataframe(
                        filtered_df,
                        column_config={
                            'Material': st.column_config.TextColumn('Material', width=180),
                            'NSOH': st.column_config.TextColumn('NSOH', width=80),
                            'AMC': st.column_config.TextColumn('AMC', width=70),
                            'PMOS': st.column_config.TextColumn('PMOS', width=70, help="Pipeline Months of Stock"),
                            'NMOS': st.column_config.NumberColumn('NMOS', width=70, format="%.2f"),
                            'TMOS': st.column_config.NumberColumn('TMOS', width=70, format="%.2f"),
                            'Identified Problem': st.column_config.TextColumn('Identified Problem', width=220),
                            'Action Point': st.column_config.TextColumn('Action Point', width=400),
                            'Responsible Body': st.column_config.TextColumn('Responsible Body', width=180),
                            'Due Date': st.column_config.TextColumn('Due Date', width=110)
                        },
                        use_container_width=True,
                        hide_index=True,
                        height=min(600, (len(filtered_df) + 1) * 45)
                    )

                    col_download1, col_download2 = st.columns(2)
                    with col_download1:
                        st.download_button(
                            label="📥 Download Filtered View (CSV)",
                            data=filtered_df.to_csv(index=False),
                            file_name=f"action_plan_filtered_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv",
                            use_container_width=True
                        )
                    with col_download2:
                        st.download_button(
                            label="📊 Download Full Action Plan (CSV)",
                            data=action_df.to_csv(index=False),
                            file_name=f"action_plan_full_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv",
                            use_container_width=True
                        )
                else:
                    st.success("✅ No action items identified")
                    st.balloons()
            else:
                st.success("✅ No procurement needed. All materials have TMOS ≥ 18 months.")
        else:
            st.info("TMOS, NMOS, or AMC data not available for supply planning")

# ---------------------------------------------------
# TAB 6 - Purchase Order Status
# ---------------------------------------------------
with tab6:
    st.markdown("<h3 style='font-size: 28px; font-weight: bold; font-family: Times New Roman;'>📋 Purchase Order Status Tracking</h3>", unsafe_allow_html=True)
    st.caption("Pipeline Purchase Orders - Track all active POs with their status and pipeline months of stock")

    if not df_filtered.empty:
        po_records = []

        for idx, row in df_filtered.iterrows():
            material = row.get('Material Description', '')
            if pd.isna(material) or material == '':
                continue

            # Get the actual Status from the Stock Status Table
            actual_status = row.get('Status', '')
            if pd.isna(actual_status) or actual_status == '':
                actual_status = 'No Status'

            # Get pipeline MOS values
            git_mos = row.get('GIT_MOS', 0)
            lc_mos = row.get('LC_MOS', 0)
            wb_mos = row.get('WB_MOS', 0)
            tmd_mos = row.get('TMD_MOS', 0)

            # Get PO numbers
            git_po = row.get('GIT_PO', '')
            lc_po = row.get('LC_PO', '')
            wb_po = row.get('WB_PO', '')
            tmd_po = row.get('TMD_PO', '')

            # Get quantities
            git_qty = row.get('GIT_Qty', 0)
            lc_qty = row.get('LC_Qty', 0)
            wb_qty = row.get('WB_Qty', 0)
            tmd_qty = row.get('TMD_Qty', 0)

            # Convert to numeric
            try:
                git_mos = float(git_mos) if pd.notna(git_mos) else 0
                lc_mos = float(lc_mos) if pd.notna(lc_mos) else 0
                wb_mos = float(wb_mos) if pd.notna(wb_mos) else 0
                tmd_mos = float(tmd_mos) if pd.notna(tmd_mos) else 0

                git_qty = float(git_qty) if pd.notna(git_qty) else 0
                lc_qty = float(lc_qty) if pd.notna(lc_qty) else 0
                wb_qty = float(wb_qty) if pd.notna(wb_qty) else 0
                tmd_qty = float(tmd_qty) if pd.notna(tmd_qty) else 0
            except:
                continue

            # Process GIT PO
            if git_mos > 0 and git_po != '' and str(git_po) != 'nan':
                po_records.append({
                    'Material Description': material,
                    'PO Number': str(git_po),
                    'Quantity': format_number_with_commas(git_qty) if git_qty > 0 else 'N/A',
                    'PMOS': round(git_mos, 2),
                    'Status': actual_status,  # CHANGED: Using actual Status from Stock Status Table
                    'Status Type': 'GIT'
                })

            # Process LC PO
            if lc_mos > 0 and lc_po != '' and str(lc_po) != 'nan':
                po_records.append({
                    'Material Description': material,
                    'PO Number': str(lc_po),
                    'Quantity': format_number_with_commas(lc_qty) if lc_qty > 0 else 'N/A',
                    'PMOS': round(lc_mos, 2),
                    'Status': actual_status,  # CHANGED: Using actual Status from Stock Status Table
                    'Status Type': 'LC'
                })

            # Process WB PO
            if wb_mos > 0 and wb_po != '' and str(wb_po) != 'nan':
                po_records.append({
                    'Material Description': material,
                    'PO Number': str(wb_po),
                    'Quantity': format_number_with_commas(wb_qty) if wb_qty > 0 else 'N/A',
                    'PMOS': round(wb_mos, 2),
                    'Status': actual_status,  # CHANGED: Using actual Status from Stock Status Table
                    'Status Type': 'WB'
                })

            # Process TMD PO
            if tmd_mos > 0 and tmd_po != '' and str(tmd_po) != 'nan':
                po_records.append({
                    'Material Description': material,
                    'PO Number': str(tmd_po),
                    'Quantity': format_number_with_commas(tmd_qty) if tmd_qty > 0 else 'N/A',
                    'PMOS': round(tmd_mos, 2),
                    'Status': actual_status,  # CHANGED: Using actual Status from Stock Status Table
                    'Status Type': 'TMD'
                })

        if po_records:
            po_df = pd.DataFrame(po_records)

            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("📋 Total POs", len(po_df))
            with col2:
                unique_materials = po_df['Material Description'].nunique()
                st.metric("📦 Materials with POs", unique_materials)
            with col3:
                total_pmos = po_df['PMOS'].sum()
                st.metric("📊 Total Pipeline MOS", f"{total_pmos:.2f}")
            with col4:
                status_counts = po_df['Status Type'].value_counts()
                most_common = status_counts.index[0] if len(status_counts) > 0 else 'N/A'
                st.metric("🔝 Most Common", most_common)

            st.markdown("---")

            # Status filter
            status_options = ["All"] + sorted(po_df['Status'].unique().tolist())
            po_status_filter = st.selectbox("Filter by Status", status_options, key="po_status_filter")

            if po_status_filter != "All":
                filtered_po_df = po_df[po_df['Status'] == po_status_filter]
                st.info(f"Showing {len(filtered_po_df)} POs with status: {po_status_filter}")
            else:
                filtered_po_df = po_df
                st.info(f"Showing all {len(filtered_po_df)} POs")

            # Search box
            po_search = st.text_input("🔍 Search by Material or PO Number", placeholder="e.g., 'artesunate' or 'PO-12345'")
            if po_search:
                search_mask = (
                    filtered_po_df['Material Description'].str.contains(po_search, case=False, na=False) |
                    filtered_po_df['PO Number'].str.contains(po_search, case=False, na=False)
                )
                filtered_po_df = filtered_po_df[search_mask]
                st.info(f"Found {len(filtered_po_df)} matching POs")

            # Display table
            st.dataframe(
                filtered_po_df,
                column_config={
                    'Material Description': st.column_config.TextColumn('Material Description', width=300),
                    'PO Number': st.column_config.TextColumn('PO Number', width=150),
                    'Quantity': st.column_config.TextColumn('Quantity', width=100),
                    'PMOS': st.column_config.NumberColumn('PMOS (Months)', width=100, format="%.2f"),
                    'Status': st.column_config.TextColumn('Status', width=200),
                    'Status Type': st.column_config.TextColumn('Type', width=80)
                },
                use_container_width=True,
                hide_index=True
            )

# ---------------------------------------------------
# TAB 7 - New Deliveries
# ---------------------------------------------------
with tab7:
    # Filter by selected program
    if not df_new_deliveries.empty and sheet_name != "All" and sheet_name in google_sheets:
        program_materials = google_sheets[sheet_name]['Material Description'].dropna().tolist()
        df_new_deliveries = df_new_deliveries[df_new_deliveries['Material Description'].isin(program_materials)]

    st.markdown("<h3 style='font-size: 28px; font-weight: bold; font-family: Times New Roman;'>🚚 New Deliveries Tracking</h3>", unsafe_allow_html=True)
    st.caption("Track incoming deliveries - Purchase Orders and Quantities")

    if not df_new_deliveries.empty:
        # Convert Posting Date to datetime
        df_new_deliveries['Posting Date'] = pd.to_datetime(df_new_deliveries['Posting Date'], errors='coerce')

        # Date range filter
        st.markdown("#### 📅 Date Range Filter")
        col_date1, col_date2 = st.columns(2)

        # Get min and max dates from data
        min_date = df_new_deliveries['Posting Date'].min().date() if not df_new_deliveries['Posting Date'].isna().all() else datetime.now().date()
        max_date = df_new_deliveries['Posting Date'].max().date() if not df_new_deliveries['Posting Date'].isna().all() else datetime.now().date()

        with col_date1:
            start_date = st.date_input("From Date", value=min_date, min_value=min_date, max_value=max_date, key="start_date")
        with col_date2:
            end_date = st.date_input("To Date", value=max_date, min_value=min_date, max_value=max_date, key="end_date")

        # Filter by date range
        mask_date = (df_new_deliveries['Posting Date'].dt.date >= start_date) & (df_new_deliveries['Posting Date'].dt.date <= end_date)
        filtered_delivery_df = df_new_deliveries[mask_date].copy()

        st.markdown("---")

        # Summary metrics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("📦 Total Deliveries", len(filtered_delivery_df))
        with col2:
            unique_pos = filtered_delivery_df['PO Number'].nunique() if 'PO Number' in filtered_delivery_df.columns else 0
            st.metric("📋 Unique POs", unique_pos)
        with col3:
            if not filtered_delivery_df.empty and 'Posting Date' in filtered_delivery_df.columns:
                newest_row = filtered_delivery_df.loc[filtered_delivery_df['Posting Date'].idxmax()]
                newest_po = newest_row.get('PO Number', 'N/A')
                newest_date = newest_row.get('Posting Date', datetime.now()).strftime('%d/%m/%y')
                st.metric("🆕 Newest PO", newest_po, delta=f"Date: {newest_date}")
            else:
                st.metric("🆕 Newest PO", "N/A")

        st.markdown("---")

        # Search filter
        search_delivery = st.text_input("🔍 Search by Material, PO Number", placeholder="e.g., 'artesunate' or 'PO-12345'")

        if search_delivery:
            mask = filtered_delivery_df.astype(str).apply(
                lambda col: col.str.contains(search_delivery, case=False, na=False)
            ).any(axis=1)
            filtered_delivery_df = filtered_delivery_df[mask]
            st.info(f"Found {len(filtered_delivery_df)} matching deliveries")

        # Format Quantity column for display
        display_df = filtered_delivery_df.copy()
        if 'Quantity' in display_df.columns:
            display_df['Quantity'] = pd.to_numeric(display_df['Quantity'], errors='coerce')
            display_df['Quantity_display'] = display_df['Quantity'].apply(format_number_with_commas)
        else:
            display_df['Quantity_display'] = 'N/A'

        # Format Posting Date column
        if 'Posting Date' in display_df.columns:
            display_df['Posting Date'] = display_df['Posting Date'].dt.strftime('%Y-%m-%d')

        # ========== TABLE 1: Detailed View (SORTED DESCENDING) ==========
        st.markdown("### 📋 Table 1: Detailed Deliveries")

        # Sort by Posting Date in descending order (newest first)
        if 'Posting Date' in display_df.columns:
            display_df = display_df.sort_values('Posting Date', ascending=False)

        # Select columns for display
        display_cols = ['Material Description', 'Posting Date', 'PO Number', 'Quantity_display']
        display_cols = [col for col in display_cols if col in display_df.columns]
        display_rename = {'Quantity_display': 'Quantity'}

        st.dataframe(
            display_df[display_cols].rename(columns=display_rename),
            column_config={
                'Material Description': st.column_config.TextColumn('Material Description', width=300),
                'Posting Date': st.column_config.TextColumn('Posting Date', width=120),
                'PO Number': st.column_config.TextColumn('PO Number', width=150),
                'Quantity': st.column_config.TextColumn('Quantity', width=100)
            },
            use_container_width=True,
            hide_index=True
        )

        # ========== TABLE 2: Summarized View ==========
        st.markdown("---")
        st.markdown("### 📊 Table 2: Summarized Deliveries (Same Material + PO Combined)")
        st.caption("When the same Material Description and PO Number appear multiple times, quantities are summed up")

        summary_df = filtered_delivery_df.copy()
        if 'Quantity' in summary_df.columns:
            summary_df['Quantity'] = pd.to_numeric(summary_df['Quantity'], errors='coerce')

        grouped_summary = summary_df.groupby(['Material Description', 'PO Number']).agg({
            'Quantity': 'sum',
            'Posting Date': 'count'
        }).rename(columns={'Quantity': 'Total Quantity', 'Posting Date': 'Number of Deliveries'}).reset_index()

        grouped_summary['Total Quantity'] = grouped_summary['Total Quantity'].apply(format_number_with_commas)

        st.dataframe(
            grouped_summary,
            column_config={
                'Material Description': st.column_config.TextColumn('Material Description', width=350),
                'PO Number': st.column_config.TextColumn('PO Number', width=180),
                'Total Quantity': st.column_config.TextColumn('Total Quantity', width=120),
                'Number of Deliveries': st.column_config.NumberColumn('Number of Deliveries', width=120)
            },
            use_container_width=True,
            hide_index=True
        )

        # ========== TOP NEW DELIVERIES ==========
        st.markdown("---")
        st.markdown("### 🏆 Top New Deliveries (Highest Quantity)")

        top_deliveries = filtered_delivery_df.copy()
        if 'Quantity' in top_deliveries.columns:
            top_deliveries['Quantity'] = pd.to_numeric(top_deliveries['Quantity'], errors='coerce')

        top_grouped = top_deliveries.groupby('Material Description').agg({
            'Quantity': 'sum'
        }).reset_index().sort_values('Quantity', ascending=False).head(10)

        top_grouped['Quantity'] = top_grouped['Quantity'].apply(format_number_with_commas)

        col_top1, col_top2, col_top3, col_top4, col_top5 = st.columns(5)
        for i, (idx, row) in enumerate(top_grouped.head(5).iterrows()):
            with [col_top1, col_top2, col_top3, col_top4, col_top5][i]:
                st.metric(f"#{i+1}", row['Material Description'][:30], delta=row['Quantity'])

        st.dataframe(
            top_grouped,
            column_config={
                'Material Description': st.column_config.TextColumn('Material Description', width=400),
                'Quantity': st.column_config.TextColumn('Total Quantity', width=150)
            },
            use_container_width=True,
            hide_index=True
        )

        # ========== NEWEST POs Section ==========
        st.markdown("---")
        st.markdown("### 🆕 Newest Purchase Orders (Last 10)")

        newest_pos = filtered_delivery_df.copy()
        if 'Posting Date' in newest_pos.columns:
            newest_pos['Posting Date'] = pd.to_datetime(newest_pos['Posting Date'], errors='coerce')
            newest_pos = newest_pos.sort_values('Posting Date', ascending=False).drop_duplicates(subset=['PO Number']).head(10)

            newest_pos['Posting Date'] = newest_pos['Posting Date'].dt.strftime('%Y-%m-%d')
            if 'Quantity' in newest_pos.columns:
                newest_pos['Quantity'] = pd.to_numeric(newest_pos['Quantity'], errors='coerce')
                newest_pos['Quantity'] = newest_pos['Quantity'].apply(format_number_with_commas)

            newest_display_cols = ['PO Number', 'Posting Date', 'Material Description', 'Quantity']
            newest_display_cols = [col for col in newest_display_cols if col in newest_pos.columns]

            st.dataframe(
                newest_pos[newest_display_cols],
                column_config={
                    'PO Number': st.column_config.TextColumn('PO Number', width=150),
                    'Posting Date': st.column_config.TextColumn('Posting Date', width=120),
                    'Material Description': st.column_config.TextColumn('Material Description', width=300),
                    'Quantity': st.column_config.TextColumn('Quantity', width=100)
                },
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("No posting date data available to determine newest POs")

        # Download buttons
        st.markdown("---")
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                label="📥 Download Detailed View (CSV)",
                data=display_df[display_cols].rename(columns=display_rename).to_csv(index=False),
                file_name=f"new_deliveries_detailed_{start_date}_{end_date}.csv",
                mime="text/csv",
                use_container_width=True
            )
        with col_dl2:
            st.download_button(
                label="📥 Download Summarized View (CSV)",
                data=grouped_summary.to_csv(index=False),
                file_name=f"new_deliveries_summarized_{start_date}_{end_date}.csv",
                mime="text/csv",
                use_container_width=True
            )

        if len(filtered_delivery_df) > len(grouped_summary):
            st.info(f"📊 {len(filtered_delivery_df)} detailed records combined into {len(grouped_summary)} summarized records")
    else:
        st.info("No new deliveries data available. Please upload data to the New_deliveries table in Supabase.")
        st.caption("Expected columns: material_description, posting_date, purchase_order, quantity")                      

# ---------------------------------------------------
# Download Filtered Data
# ---------------------------------------------------
if not display_df_filtered.empty and 'Material Description' in display_df_filtered.columns and page != "Advanced Analytics" and page != "Executive Summary":
    st.divider()
    st.download_button(label="📥 Download Full Data (CSV)", data=display_df_filtered.to_csv(index=False), 
                      file_name=f"full_stock_dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv", 
                      mime="text/csv", use_container_width=True)
